import streamlit as st
import pandas as pd
import io
import os
import math
import json
import calendar
import hashlib
import zipfile
from fpdf import FPDF
from datetime import datetime, timedelta, date
from pathlib import Path
from decimal import Decimal, ROUND_HALF_UP, InvalidOperation
import bisect
import sys

from typing import Any, Callable


PDF_AUTHOR_NAME = "Er.Aravind MRT VREDC"


def _robust_to_datetime(series):
    """Try parsing a Series of dates flexibly: first day-first, then month-first.
    Returns a pd.Series of datetimes (may contain NaT where parsing fails).
    """
    try:
        parsed = pd.to_datetime(series, errors='coerce', dayfirst=True)
        if parsed.notna().any():
            return parsed
        # Fallback: try without dayfirst
        parsed = pd.to_datetime(series, errors='coerce', dayfirst=False)
        return parsed
    except Exception:
        return pd.to_datetime(series, errors='coerce')

# Additional Surcharge configuration (rates vary by regulatory period)
ADDITIONAL_SURCHARGE_WINDOWS = [
    {
        "start": datetime(2021, 4, 16),
        "end": datetime(2022, 3, 31),
        "rate": 0.70,
        "note": "TNERC M.P.No.18 of 2020 dt.15.04.2021",
    },
    {
        "start": datetime(2023, 2, 25),
        "end": datetime(2023, 3, 31),
        "rate": 0.83,
        "note": "TNERC M.P.No.32 of 2021 dt.08.02.2022",
    },
    {
        "start": datetime(2024, 12, 12),
        "end": datetime(2025, 3, 31),
        "rate": 0.54,
        "note": "TNERC M.P.No.44 of 2024 dt.12.12.2024",
    },
    {
        "start": datetime(2025, 4, 29),
        "end": datetime(2025, 9, 30),
        "rate": 0.10,
        "note": "TNERC M.P.No.13 of 2025 dt.29.04.2025",
    },
]


def _get_month_period_bounds(month_value, year_value):
    """Return (month_start, month_end, label) for given month/year strings."""
    if not month_value or not year_value:
        return None
    try:
        month_int = int(float(month_value))
        year_int = int(float(year_value))
        if month_int < 1 or month_int > 12:
            return None
        last_day = calendar.monthrange(year_int, month_int)[1]
        period_start = datetime(year_int, month_int, 1)
        period_end = datetime(year_int, month_int, last_day)
        period_label = f"{calendar.month_name[month_int]} {year_int}"
        return period_start, period_end, period_label
    except Exception:
        return None


def _resolve_additional_surcharge_rate(month_value, year_value):
    """Determine applicable rate/note for the selected month/year."""
    bounds = _get_month_period_bounds(month_value, year_value)
    if not bounds:
        return 0.0, "Select a valid month/year to apply Additional Surcharge", "Month & Year not selected", "Month & Year not selected"
    period_start, period_end, period_label = bounds
    for window in ADDITIONAL_SURCHARGE_WINDOWS:
        if period_end >= window["start"] and period_start <= window["end"]:
            window_label = f"{window['start'].strftime('%d-%m-%Y')} to {window['end'].strftime('%d-%m-%Y')}"
            return window["rate"], window["note"], period_label, window_label
    return 0.0, f"No Additional Surcharge defined for {period_label}", period_label, "No matching TNERC window"


def calculate_monthly_additional_surcharge(month_value, year_value, iex_excess_raw, iex_excess_rounded):
    """Compute Additional Surcharge (IEX) based on selected month/year and IEX excess energy."""
    rate, note, period_label, window_label = _resolve_additional_surcharge_rate(month_value, year_value)
    iex_excess_raw = iex_excess_raw or 0.0
    iex_excess_rounded = iex_excess_rounded or 0
    component = iex_excess_raw * rate if rate and iex_excess_raw else 0.0
    breakdown_entry = {
        "period_label": period_label,
        "window_label": window_label,
        "kwh": iex_excess_rounded,
        "raw_kwh": iex_excess_raw,
        "rate": rate,
        "amount": component,
        "note": note,
    }
    breakdown = [breakdown_entry]
    return component, breakdown, rate, period_label, note


TARIFF_OPTIONS = ["Tariff I", "Tariff II-A", "Tariff II-B", "Tariff III"]

TARIFF_WINDOWS = [
    {
        "start": datetime(2022, 9, 10),
        "label": "W.E.F 10.09.2022",
        "rates": {
            "Tariff I": {"base_rate": 6.75, "c1_c2_rate": 1.6875, "c5_rate": 0.3375, "wheeling_rate": 0.96, "cross_subsidy_rate": 1.79},
            "Tariff II-A": {"base_rate": 7.00, "c1_c2_rate": 1.75, "c5_rate": 0.35, "wheeling_rate": 0.96, "cross_subsidy_rate": 1.95},
            "Tariff II-B": {"base_rate": 7.50, "c1_c2_rate": 1.88, "c5_rate": 0.375, "wheeling_rate": 0.96, "cross_subsidy_rate": 2.19},
            "Tariff III": {"base_rate": 7.50, "c1_c2_rate": 1.88, "c5_rate": 0.425, "wheeling_rate": 0.96, "cross_subsidy_rate": 2.33},
        },
    },
    {
        "start": datetime(2023, 7, 1),
        "label": "W.E.F 01.07.2023",
        "rates": {
            "Tariff I": {"base_rate": 6.90, "c1_c2_rate": 1.725, "c5_rate": 0.345, "wheeling_rate": 1.0, "cross_subsidy_rate": 1.86},
            "Tariff II-A": {"base_rate": 7.15, "c1_c2_rate": 1.7875, "c5_rate": 0.3575, "wheeling_rate": 1.0, "cross_subsidy_rate": 2.02},
            "Tariff II-B": {"base_rate": 7.65, "c1_c2_rate": 1.91, "c5_rate": 0.383, "wheeling_rate": 1.0, "cross_subsidy_rate": 2.27},
            "Tariff III": {"base_rate": 8.70, "c1_c2_rate": 2.175, "c5_rate": 0.435, "wheeling_rate": 1.0, "cross_subsidy_rate": 2.41},
        },
    },
    {
        "start": datetime(2024, 7, 1),
        "label": "W.E.F 01.07.2024",
        "rates": {
            "Tariff I": {"base_rate": 7.25, "c1_c2_rate": 1.8125, "c5_rate": 0.3625, "wheeling_rate": 1.04, "cross_subsidy_rate": 1.92},
            "Tariff II-A": {"base_rate": 7.50, "c1_c2_rate": 1.875, "c5_rate": 0.375, "wheeling_rate": 1.04, "cross_subsidy_rate": 2.11},
            "Tariff II-B": {"base_rate": 8.00, "c1_c2_rate": 2.0, "c5_rate": 0.4, "wheeling_rate": 1.04, "cross_subsidy_rate": 2.36},
            "Tariff III": {"base_rate": 9.10, "c1_c2_rate": 2.275, "c5_rate": 0.455, "wheeling_rate": 1.04, "cross_subsidy_rate": 2.49},
        },
    },
    {
        "start": datetime(2025, 7, 1),
        "label": "W.E.F 01.07.2025",
        "rates": {
            "Tariff I": {"base_rate": 7.50, "c1_c2_rate": 1.875, "c5_rate": 0.375, "wheeling_rate": 1.04, "cross_subsidy_rate": 1.99},
            "Tariff II-A": {"base_rate": 7.75, "c1_c2_rate": 1.9375, "c5_rate": 0.3875, "wheeling_rate": 1.04, "cross_subsidy_rate": 1.99},
            "Tariff II-B": {"base_rate": 8.25, "c1_c2_rate": 2.06, "c5_rate": 0.413, "wheeling_rate": 1.04, "cross_subsidy_rate": 2.46},
            "Tariff III": {"base_rate": 9.40, "c1_c2_rate": 2.35, "c5_rate": 0.47, "wheeling_rate": 1.04, "cross_subsidy_rate": 2.57},
        },
    },
]


WHEELING_PERCENT = 0.0234


def _round_kwh_half_up(value):
    try:
        value = float(value)
    except Exception:
        value = 0.0
    return int(value + 0.5) if value >= 0 else int(value - 0.5)


def compute_wheeling_components(total_excess_kwh, t_and_d_loss_percent):
    """Return wheeling helper values per revised TN formula."""
    try:
        loss_pct = float(t_and_d_loss_percent or 0)
    except Exception:
        loss_pct = 0.0

    reference_raw = 0.0
    if loss_pct > 0 and loss_pct < 100:
        reference_raw = (total_excess_kwh * loss_pct) / (100 - loss_pct)

    reference_kwh = _round_kwh_half_up(reference_raw)
    combined_kwh = total_excess_kwh + reference_kwh
    reduction_component = combined_kwh * WHEELING_PERCENT
    adjusted_kwh = combined_kwh - reduction_component

    return {
        "reference_raw": reference_raw,
        "reference_kwh": reference_kwh,
        "combined_kwh": combined_kwh,
        "reduction_kwh": reduction_component,
        "adjusted_kwh": adjusted_kwh,
    }


class AuthorPDF(FPDF):
    def __init__(self, author_name=PDF_AUTHOR_NAME, *args, **kwargs):
        super().__init__(*args, **kwargs)
        self.author_name = author_name

    def header(self):
        if self.author_name:
            self.set_font('Arial', 'I', 10)
            self.set_text_color(80, 80, 80)
            self.cell(0, 10, self.author_name, 0, 0, 'R')
            self.ln(5)
            self.set_text_color(0, 0, 0)


def _resolve_tariff_window(target_date: datetime):
    windows = sorted(TARIFF_WINDOWS, key=lambda w: w["start"])
    for idx, window in enumerate(windows):
        start = window["start"]
        if idx + 1 < len(windows):
            end = windows[idx + 1]["start"] - timedelta(days=1)
        else:
            end = datetime(2999, 12, 31)
        if target_date >= start and target_date <= end:
            return window, start, end
    return windows[-1], windows[-1]["start"], datetime(2999, 12, 31)


def resolve_tariff_rates(tariff_choice, month_value, year_value):
    bounds = _get_month_period_bounds(month_value, year_value)
    if bounds:
        period_start, _, period_label = bounds
        reference_date = period_start
    else:
        reference_date = datetime.today()
        period_label = "Month & Year not selected"

    window, _, _ = _resolve_tariff_window(reference_date)
    selected_tariff = tariff_choice if tariff_choice in TARIFF_OPTIONS else TARIFF_OPTIONS[0]
    rates = window["rates"].get(selected_tariff, window["rates"][TARIFF_OPTIONS[0]]).copy()
    rates.update({
        "tariff": selected_tariff,
        "window_label": window["label"],
        "period_label": period_label,
    })
    return rates

# Default no-op updater stub and callables. We define these first and then
# attempt to overwrite them with real implementations if `auto_updater` is
# available. Defining them first avoids redeclaration warnings from static
# analyzers.
class _UpdaterStub:
    def __init__(self) -> None:
        self.config = {}

    def check_for_updates(self, *args, **kwargs):
        return None

    def download_update(self, *args, **kwargs):
        return None

    def apply_update(self, *args, **kwargs):
        return False

# Placeholders for the real implementations (if available). We keep these
# internal so that the public API functions below have a single declaration
# (avoids redeclaration diagnostics from static analyzers).
_initialize_updater_impl: Callable[..., Any] = None  # type: ignore
_manual_update_check_impl: Callable[..., Any] = None  # type: ignore
_show_update_settings_impl: Callable[..., Any] = None  # type: ignore
UPDATER_AVAILABLE = False

# If the optional `auto_updater` package is installed, import into internal
# names and mark UPDATER_AVAILABLE. Use temporary names and cast to Any to
# avoid strict signature/type mismatch diagnostics.
try:
    from auto_updater import (
        initialize_updater as _real_initialize_updater,
        manual_update_check as _real_manual_update_check,
        show_update_settings as _real_show_update_settings,
    )
    _initialize_updater_impl = _real_initialize_updater  # type: Any
    _manual_update_check_impl = _real_manual_update_check  # type: Any
    _show_update_settings_impl = _real_show_update_settings  # type: Any
    UPDATER_AVAILABLE = True
except Exception:
    # Keep the internal impls as None; we will fall back to stubs in wrappers
    UPDATER_AVAILABLE = False

# Public wrapper functions (single declaration each) that delegate to the
# real implementations when available, otherwise use safe fallbacks.
def initialize_updater(current_version: str = "1.0.0") -> Any:
    if _initialize_updater_impl is not None:
        return _initialize_updater_impl(current_version)
    return _UpdaterStub()


def manual_update_check(*args, **kwargs) -> Any:
    if _manual_update_check_impl is not None:
        return _manual_update_check_impl(*args, **kwargs)
    return None


def show_update_settings(*args, **kwargs) -> Any:
    if _show_update_settings_impl is not None:
        return _show_update_settings_impl(*args, **kwargs)
    return None

# Set page config
st.set_page_config(
    page_title="Energy Adjustment Calculator",
    page_icon="⚡",
    layout="wide"
)


def render_auto_update_sidebar() -> None:
    """Render the auto-update UI in the Streamlit sidebar (when available)."""
    if not UPDATER_AVAILABLE:
        return

    with st.sidebar:
        st.header("🔄 Auto-Update System")

        # Get current version
        version_file = Path(__file__).parent / "version.json"
        current_version = "1.0.0"
        if version_file.exists():
            try:
                with open(version_file, 'r') as f:
                    version_data = json.load(f)
                    current_version = version_data.get("version", "1.0.0")
            except Exception:
                pass

        st.info(f"**Current Version:** {current_version}")

        # Initialize updater
        if 'updater' not in st.session_state:
            st.session_state.updater = initialize_updater(current_version)

        col1, col2 = st.columns(2)

        with col1:
            if st.button("🔍 Check Updates", help="Check for available updates"):
                with st.spinner("Checking for updates..."):
                    update_info = st.session_state.updater.check_for_updates(show_no_updates=False)
                    if update_info:
                        st.success(f"Update available: v{update_info['version']}")
                        st.info(f"**What's New:**\n{update_info['description'][:200]}...")

                        if st.button("📥 Download Update", key="download_update"):
                            with st.spinner("Downloading and applying update..."):
                                try:
                                    update_file = st.session_state.updater.download_update(update_info)
                                    if update_file:
                                        if st.session_state.updater.apply_update(update_file, update_info):
                                            st.success("✅ Update applied successfully! Please restart the application.")
                                            st.balloons()
                                        else:
                                            st.error("❌ Failed to apply update.")
                                    else:
                                        st.error("❌ Failed to download update.")
                                except Exception as e:
                                    st.error(f"❌ Update failed: {e}")
                    else:
                        st.success("✅ You have the latest version!")

        with col2:
            if st.button("⚙️ Settings", help="Update settings"):
                show_update_settings(st.session_state.updater)

        # Auto-update status
        if st.session_state.updater.config.get("auto_check", True):
            st.success("🟢 Auto-updates enabled")
        else:
            st.warning("🟡 Auto-updates disabled")

        # Last check info
        last_check = st.session_state.updater.config.get("last_check")
        if last_check:
            try:
                last_check_time = datetime.fromisoformat(last_check)
                st.caption(f"Last checked: {last_check_time.strftime('%d/%m/%Y %H:%M')}")
            except Exception:
                pass


def render_footer(app_label: str) -> None:
    st.markdown("---")
    st.markdown(app_label)
    if UPDATER_AVAILABLE:
        st.markdown("🔄 Auto-update system enabled")


def render_bpsc_calculator() -> None:
    def _sanitize_excel_text(value: str) -> str:
        """Prevent Excel formula injection for exported strings."""
        text = (value or "")
        if not text:
            return ""
        if text[0] in ('=', '+', '-', '@'):
            return "'" + text
        return text

    def _parse_optional_day(day_text: str | None) -> int | None:
        if day_text is None:
            return None
        day_text = str(day_text).strip()
        if not day_text:
            return None
        try:
            day_val = int(float(day_text))
        except Exception:
            return None
        if day_val < 1 or day_val > 31:
            return None
        return day_val

    def _build_date_from_month_year_optional_day(month_int: int, year_int: int, day_opt: int | None) -> datetime:
        last_day = calendar.monthrange(year_int, month_int)[1]
        day_int = day_opt if day_opt is not None else last_day
        if day_int > last_day:
            day_int = last_day
        return datetime(year_int, month_int, day_int)

    def _render_month_year_day(
        prefix: str,
        label: str,
        default_month_int: int | None = None,
        default_year_int: int | None = None,
        disable_month_year: bool = False,
    ) -> tuple[datetime, str, int, int]:
        now = datetime.today()
        months = list(range(1, 13))
        month_labels = [calendar.month_name[m] for m in months]

        if default_month_int is None:
            default_month_int = now.month
        if default_year_int is None:
            default_year_int = now.year

        default_month_int = int(default_month_int)
        if default_month_int < 1 or default_month_int > 12:
            default_month_int = now.month

        years = list(range(2000, 2101))
        if int(default_year_int) not in years:
            default_year_int = now.year

        colm, coly, cold = st.columns([2, 2, 1])
        with colm:
            month_idx = st.selectbox(
                f"{label} Month",
                options=list(range(len(months))),
                format_func=lambda i: month_labels[i],
                index=max(0, int(default_month_int) - 1),
                key=f"{prefix}_month_idx",
                disabled=disable_month_year,
            )
            month_int = months[month_idx]
        with coly:
            year_int = st.selectbox(
                f"{label} Year",
                options=years,
                index=years.index(int(default_year_int)),
                key=f"{prefix}_year",
                disabled=disable_month_year,
            )
        with cold:
            day_text = st.text_input(
                f"{label} Day (optional)",
                value="",
                key=f"{prefix}_day",
                help="Leave blank to default to the last day of the selected month.",
            )

        day_opt = _parse_optional_day(day_text)
        dt = _build_date_from_month_year_optional_day(month_int, year_int, day_opt)
        display = dt.strftime('%d/%m/%Y')
        return dt, display, month_int, year_int

    def _generate_bpsc_pdf(
        consumer_name: str,
        service_number: str,
        author_name: str,
        rows: list[dict[str, Any]],
        totals: dict[str, float],
    ) -> bytes:
        pdf = AuthorPDF(author_name=author_name)
        pdf.set_auto_page_break(auto=True, margin=15)
        pdf.add_page()

        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'BPSC Calculator Report', ln=True, align='C')
        pdf.ln(2)

        pdf.set_font('Arial', '', 11)
        pdf.cell(0, 7, f"Consumer Name: {consumer_name}", ln=True)
        pdf.cell(0, 7, f"Service Number: {service_number}", ln=True)
        pdf.ln(4)

        headers = ['S.No', 'Amount', 'Bill Issued', 'Due Date', 'Cutoff Date', 'Days', 'BPSC Amount', 'Total']
        col_widths = [10, 22, 24, 24, 24, 12, 26, 28]

        pdf.set_font('Arial', 'B', 10)
        for h, w in zip(headers, col_widths):
            pdf.cell(w, 8, h, border=1, align='C')
        pdf.ln()

        pdf.set_font('Arial', '', 10)
        for r in rows:
            pdf.cell(col_widths[0], 8, str(r['sno']), border=1, align='C')
            pdf.cell(col_widths[1], 8, f"{r['amount']:.2f}", border=1, align='R')
            pdf.cell(col_widths[2], 8, str(r.get('issued_date', '')), border=1, align='C')
            pdf.cell(col_widths[3], 8, r['due_date'], border=1, align='C')
            pdf.cell(col_widths[4], 8, r['cutoff_date'], border=1, align='C')
            pdf.cell(col_widths[5], 8, str(r['days']), border=1, align='C')
            pdf.cell(col_widths[6], 8, f"{r['bpsc_amount']:.2f}", border=1, align='R')
            pdf.cell(col_widths[7], 8, f"{r['row_total']:.2f}", border=1, align='R')
            pdf.ln()

        pdf.ln(4)
        pdf.set_font('Arial', 'B', 11)
        pdf.cell(0, 7, f"Total Base Amount: {totals['total_base']:.2f}", ln=True)
        pdf.cell(0, 7, f"Total BPSC: {totals['total_bpsc']:.2f}", ln=True)
        pdf.cell(0, 7, f"Final Amount: {totals['final_amount']:.2f}", ln=True)

        return pdf.output(dest='S').encode('latin-1')

    def _generate_bpsc_excel(
        consumer_name: str,
        service_number: str,
        author_name: str,
        rows: list[dict[str, Any]],
        totals: dict[str, float],
    ) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "BPSC Report"

        bold = Font(bold=True)
        title_font = Font(bold=True, size=14)
        thin = Side(style='thin')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        ws.merge_cells('A1:H1')
        ws['A1'] = 'BPSC Calculator Report'
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')

        ws['A3'] = 'Author:'
        ws['B3'] = _sanitize_excel_text(author_name)
        ws['A4'] = 'Consumer Name:'
        ws['B4'] = _sanitize_excel_text(consumer_name)
        ws['A5'] = 'Service Number:'
        ws['B5'] = _sanitize_excel_text(service_number)
        for cell in ['A3', 'A4', 'A5']:
            ws[cell].font = bold

        start_row = 7
        headers = ['S.No', 'Amount', 'Bill Issued', 'Due Date', 'Cutoff Date', 'Difference (Days)', 'BPSC Amount', 'Base Amount + BPSC']
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = bold
            cell.alignment = Alignment(horizontal='center')
            cell.border = border

        for i, r in enumerate(rows, start=1):
            row_num = start_row + i
            values = [
                r['sno'],
                round(float(r['amount']), 2),
                _sanitize_excel_text(str(r.get('issued_date', ''))),
                _sanitize_excel_text(r['due_date']),
                _sanitize_excel_text(r['cutoff_date']),
                int(r['days']),
                round(float(r['bpsc_amount']), 2),
                round(float(r['row_total']), 2),
            ]
            for col_idx, val in enumerate(values, start=1):
                cell = ws.cell(row=row_num, column=col_idx, value=val)
                cell.border = border
                if col_idx in (1, 6):
                    cell.alignment = Alignment(horizontal='center')
                elif col_idx in (3, 4, 5):
                    cell.alignment = Alignment(horizontal='center')
                else:
                    cell.alignment = Alignment(horizontal='right')

        totals_row = start_row + len(rows) + 2
        ws['A' + str(totals_row)] = 'Total Base Amount (₹)'
        ws['B' + str(totals_row)] = float(totals['total_base'])
        ws['A' + str(totals_row + 1)] = 'Total BPSC (₹)'
        ws['B' + str(totals_row + 1)] = float(totals['total_bpsc'])
        ws['A' + str(totals_row + 2)] = 'Final Amount (₹)'
        ws['B' + str(totals_row + 2)] = float(totals['final_amount'])
        for r in range(totals_row, totals_row + 3):
            ws['A' + str(r)].font = bold

        # Basic column widths
        widths = [6, 14, 14, 14, 14, 16, 14, 18]
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[chr(64 + i)].width = w

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue()

    # Header
    col1, col2 = st.columns([3, 1])
    with col1:
        st.title("BPSC Calculator")
    with col2:
        st.markdown(f"**Author: {PDF_AUTHOR_NAME}**")

    # Session state storage for entries
    if 'bpsc_entries' not in st.session_state:
        st.session_state.bpsc_entries = []
    if 'bpsc_processed' not in st.session_state:
        st.session_state.bpsc_processed = False
    if 'bpsc_next_id' not in st.session_state:
        st.session_state.bpsc_next_id = 1
    # Do not directly mutate session_state for widget keys after instantiation.
    # We avoid form clear_on_submit for date widgets (it forces defaults back).
    if 'bpsc_clear_amount_next' not in st.session_state:
        st.session_state.bpsc_clear_amount_next = False

    # Clear amount input BEFORE the widget is created
    if st.session_state.get('bpsc_clear_amount_next'):
        st.session_state.pop('bpsc_amount_text', None)
        st.session_state.bpsc_clear_amount_next = False

    def _parse_amount_text(amount_text: str) -> float | None:
        raw = (amount_text or "").strip()
        if not raw:
            return None
        # allow copy/paste with commas (e.g., 1,23,456.78)
        cleaned = raw.replace(",", "").replace(" ", "")
        try:
            val = float(cleaned)
        except Exception:
            return None
        if val < 0:
            return None
        return val

    st.subheader("Consumer Information")
    c1, c2 = st.columns(2)
    with c1:
        consumer_name = st.text_input("Consumer Name", value=st.session_state.get('bpsc_consumer_name', ''))
    with c2:
        service_number = st.text_input("Service Number", value=st.session_state.get('bpsc_service_number', ''))

    st.session_state.bpsc_consumer_name = consumer_name
    st.session_state.bpsc_service_number = service_number

    st.subheader("BPSC Settings")
    s1, s2 = st.columns(2)
    with s1:
        use_fixed_bpsc = st.checkbox("Use fixed BPSC % for all entries", value=st.session_state.get('bpsc_use_fixed_pct', True))
    with s2:
        use_fixed_cutoff = st.checkbox("Use fixed cutoff date for all entries", value=st.session_state.get('bpsc_use_fixed_cutoff', True))

    st.session_state.bpsc_use_fixed_pct = use_fixed_bpsc
    st.session_state.bpsc_use_fixed_cutoff = use_fixed_cutoff

    fixed_bpsc_pct = None
    fixed_cutoff_dt = None
    fixed_cutoff_display = None

    if use_fixed_bpsc:
        fixed_bpsc_pct = st.number_input(
            "Fixed BPSC Percentage (%)",
            min_value=0.0,
            value=float(st.session_state.get('bpsc_fixed_pct', 0.0) or 0.0),
            step=0.01,
            format="%.4f",
        )
        st.session_state.bpsc_fixed_pct = fixed_bpsc_pct

    if use_fixed_cutoff:
        fixed_cutoff_dt, fixed_cutoff_display, _, _ = _render_month_year_day("bpsc_fixed_cutoff", "Cutoff Date")

    st.subheader("Add Entry")

    # NOTE: Do not place these widgets inside a `st.form`.
    # Widgets in a form don't update interactively until submit, which makes the
    # Due Month/Year override feel "stuck".
    amount_text = st.text_input(
        "Amount (₹)",
        key="bpsc_amount_text",
        help="You can paste values like 123456 or 1,23,456.00",
    )

    st.markdown("**Bill Issued Date**")
    issued_dt, issued_display, issued_month_int, issued_year_int = _render_month_year_day("bpsc_issued", "Bill Issued")

    # Due month/year default: immediate next month of Bill Issued.
    # If Bill Issued month is December, Due becomes January and year increments.
    due_auto = st.checkbox(
        "Auto-set Due Month/Year from Bill Issued",
        value=bool(st.session_state.get('bpsc_due_auto', True)),
        help="Defaults Due Month to the immediate next month of Bill Issued. Uncheck to override.",
    )
    st.session_state.bpsc_due_auto = due_auto

    due_default_month = issued_month_int + 1
    due_default_year = issued_year_int
    if due_default_month == 13:
        due_default_month = 1
        due_default_year = issued_year_int + 1

    st.markdown("**Due Date**")
    # Use separate widget keys for auto vs manual so toggling works cleanly.
    due_prefix = (
        f"bpsc_due_auto_{int(issued_year_int)}_{int(issued_month_int)}"
        if due_auto
        else "bpsc_due_manual"
    )
    due_dt, due_display, _, _ = _render_month_year_day(
        due_prefix,
        "Due",
        default_month_int=due_default_month,
        default_year_int=due_default_year,
        disable_month_year=bool(due_auto),
    )

    if not use_fixed_cutoff:
        st.markdown("**Cutoff Date**")
        cutoff_dt, cutoff_display, _, _ = _render_month_year_day("bpsc_cutoff", "Cutoff")
    else:
        cutoff_dt, cutoff_display = fixed_cutoff_dt, fixed_cutoff_display

    if not use_fixed_bpsc:
        entry_bpsc_pct = st.number_input(
            "BPSC Percentage (%)",
            min_value=0.0,
            value=0.0,
            step=0.01,
            format="%.4f",
        )
    else:
        entry_bpsc_pct = float(fixed_bpsc_pct or 0.0)

    add_clicked = st.button("Add", type="primary")
    if add_clicked:
        parsed_amount = _parse_amount_text(amount_text)
        if not consumer_name.strip() or not service_number.strip():
            st.error("Please enter Consumer Name and Service Number before adding entries.")
        elif parsed_amount is None:
            st.error("Please enter a valid Amount (non-negative number).")
        elif cutoff_dt is None or due_dt is None:
            st.error("Please provide valid Due Date and Cutoff Date.")
        elif cutoff_dt < due_dt:
            st.error("Cutoff Date must be on or after Due Date.")
        else:
            days = (cutoff_dt - due_dt).days + 1
            bpsc_amount = (float(parsed_amount) * float(entry_bpsc_pct) / 100.0 / 30.0) * float(days)
            row_total = float(parsed_amount) + float(bpsc_amount)
            st.session_state.bpsc_entries.append(
                {
                    "id": int(st.session_state.bpsc_next_id),
                    "amount": float(parsed_amount),
                    "bpsc_pct": float(entry_bpsc_pct),
                    "issued_date": issued_display,
                    "due_date": due_display,
                    "cutoff_date": cutoff_display,
                    "days": int(days),
                    "bpsc_amount": float(bpsc_amount),
                    "row_total": float(row_total),
                }
            )
            st.session_state.bpsc_next_id = int(st.session_state.bpsc_next_id) + 1
            st.session_state.bpsc_processed = False
            st.session_state.bpsc_clear_amount_next = True
            st.success("Entry added.")

    actions_col1, actions_col2 = st.columns([1, 2])
    with actions_col1:
        if st.button("Clear All Entries"):
            st.session_state.bpsc_entries = []
            st.session_state.bpsc_processed = False
            st.success("All entries cleared.")

    if not st.session_state.bpsc_entries:
        return

    def _parse_ddmmyyyy(date_text: str) -> datetime:
        cleaned = (date_text or '').strip().replace('-', '/')
        return datetime.strptime(cleaned, '%d/%m/%Y')

    # Ensure every entry has a stable id (migration for old session state)
    max_seen = 0
    for entry in st.session_state.bpsc_entries:
        if 'id' not in entry:
            entry['id'] = int(st.session_state.bpsc_next_id)
            st.session_state.bpsc_next_id = int(st.session_state.bpsc_next_id) + 1
        max_seen = max(max_seen, int(entry.get('id', 0)))
    if int(st.session_state.bpsc_next_id) <= max_seen:
        st.session_state.bpsc_next_id = max_seen + 1

    # Build editable view model
    fixed_cutoff_date = fixed_cutoff_dt.date() if (use_fixed_cutoff and fixed_cutoff_dt is not None) else None
    due_rows = []
    for entry in st.session_state.bpsc_entries:
        try:
            due_date = _parse_ddmmyyyy(entry.get('due_date', '01/01/2000')).date()
        except Exception:
            due_date = datetime(2000, 1, 1).date()
        try:
            cutoff_date = _parse_ddmmyyyy(entry.get('cutoff_date', '01/01/2000')).date()
        except Exception:
            cutoff_date = datetime(2000, 1, 1).date()
        if fixed_cutoff_date is not None:
            cutoff_date = fixed_cutoff_date
        due_rows.append(
            {
                "_id": int(entry['id']),
                "Amount": float(entry.get('amount', 0.0) or 0.0),
                "Due Date": due_date,
                "Cutoff Date": cutoff_date,
                "Remove": False,
            }
        )

    editor_df = pd.DataFrame(due_rows).set_index('_id')
    editor_df = editor_df.sort_values(by=["Due Date", "Cutoff Date"], ascending=True)

    st.subheader("Result")
    edited_df = st.data_editor(
        editor_df,
        hide_index=True,
        width='stretch',
        column_config={
            "Amount": st.column_config.NumberColumn(format="%.2f", step=1.0, min_value=0.0),
            "Due Date": st.column_config.DateColumn(format="DD/MM/YYYY"),
            "Cutoff Date": st.column_config.DateColumn(format="DD/MM/YYYY"),
            "Remove": st.column_config.CheckboxColumn(help="Tick to remove this row"),
        },
        disabled=["Cutoff Date"],
        key="bpsc_editor",
    )

    reprocess_clicked = st.button("Reprocess", type="secondary")
    if reprocess_clicked:
        # Apply edits back into session_state entries ONLY when explicitly requested.
        entries_by_id = {int(e['id']): e for e in st.session_state.bpsc_entries}
        new_entries = []

        fixed_pct_value = float(fixed_bpsc_pct or 0.0) if use_fixed_bpsc else None

        for row_id, row in edited_df.iterrows():
            row_id = int(row_id)
            entry = entries_by_id.get(row_id)
            if entry is None:
                continue

            if bool(row.get('Remove', False)):
                continue

            amount_val = float(row.get('Amount', 0.0) or 0.0)
            due_date_val = row.get('Due Date')
            cutoff_date_val = row.get('Cutoff Date')

            # Normalize date types (Streamlit DateColumn returns `datetime.date`)
            if isinstance(due_date_val, datetime):
                due_date_date = due_date_val.date()
            elif isinstance(due_date_val, date):
                due_date_date = due_date_val
            else:
                try:
                    due_date_date = _parse_ddmmyyyy(str(due_date_val)).date()
                except Exception:
                    due_date_date = datetime(2000, 1, 1).date()

            if fixed_cutoff_date is not None:
                cutoff_date_date = fixed_cutoff_date
            else:
                if isinstance(cutoff_date_val, datetime):
                    cutoff_date_date = cutoff_date_val.date()
                elif isinstance(cutoff_date_val, date):
                    cutoff_date_date = cutoff_date_val
                else:
                    try:
                        cutoff_date_date = _parse_ddmmyyyy(str(cutoff_date_val)).date()
                    except Exception:
                        cutoff_date_date = datetime(2000, 1, 1).date()

            entry['amount'] = amount_val
            entry['due_date'] = due_date_date.strftime('%d/%m/%Y')
            entry['cutoff_date'] = cutoff_date_date.strftime('%d/%m/%Y')

            calc_pct = float(entry.get('bpsc_pct', 0.0) or 0.0)
            if fixed_pct_value is not None:
                calc_pct = fixed_pct_value
                entry['bpsc_pct'] = calc_pct

            if cutoff_date_date < due_date_date:
                days_val = 0
            else:
                days_val = (cutoff_date_date - due_date_date).days + 1

            bpsc_amount_val = (amount_val * calc_pct / 100.0 / 30.0) * float(days_val)
            row_total_val = amount_val + bpsc_amount_val

            entry['days'] = int(days_val)
            entry['bpsc_amount'] = float(bpsc_amount_val)
            entry['row_total'] = float(row_total_val)

            new_entries.append(entry)

        def _safe_due_dt(e: dict[str, Any]) -> datetime:
            try:
                return _parse_ddmmyyyy(e.get('due_date', '01/01/2000'))
            except Exception:
                return datetime(2000, 1, 1)

        st.session_state.bpsc_entries = sorted(new_entries, key=lambda e: (_safe_due_dt(e), int(e.get('id', 0))))
        st.rerun()

    # Build output table for downloads (sorted already)
    rows_for_table: list[dict[str, Any]] = []
    for idx, entry in enumerate(st.session_state.bpsc_entries, start=1):
        rows_for_table.append(
            {
                "sno": idx,
                "amount": float(entry.get("amount", 0.0) or 0.0),
                "due_date": entry.get("due_date", ""),
                "cutoff_date": entry.get("cutoff_date", ""),
                "days": int(entry.get("days", 0) or 0),
                "bpsc_amount": float(entry.get("bpsc_amount", 0.0) or 0.0),
                "row_total": float(entry.get("row_total", 0.0) or 0.0),
                "issued_date": entry.get("issued_date", ""),
                "bpsc_pct": float(entry.get("bpsc_pct", 0.0) or 0.0),
            }
        )

    total_base = sum(r["amount"] for r in rows_for_table)
    total_bpsc = sum(r["bpsc_amount"] for r in rows_for_table)
    final_amount = total_base + total_bpsc

    totals = {
        "total_base": float(total_base),
        "total_bpsc": float(total_bpsc),
        "final_amount": float(final_amount),
    }

    m1, m2, m3 = st.columns(3)
    m1.metric("Total Base Amount (₹)", f"{total_base:,.2f}")
    m2.metric("Total BPSC (₹)", f"{total_bpsc:,.2f}")
    m3.metric("Final Amount (₹)", f"{final_amount:,.2f}")

    # Downloads
    pdf_bytes = _generate_bpsc_pdf(
        consumer_name=consumer_name,
        service_number=service_number,
        author_name=PDF_AUTHOR_NAME,
        rows=rows_for_table,
        totals=totals,
    )
    excel_bytes = _generate_bpsc_excel(
        consumer_name=consumer_name,
        service_number=service_number,
        author_name=PDF_AUTHOR_NAME,
        rows=rows_for_table,
        totals=totals,
    )

    dl1, dl2 = st.columns(2)
    with dl1:
        st.download_button(
            label="Download PDF",
            data=pdf_bytes,
            file_name=f"{service_number}_BPSC_Report.pdf" if service_number else "BPSC_Report.pdf",
            mime="application/pdf",
            type="primary",
        )
    with dl2:
        st.download_button(
            label="Download Excel",
            data=excel_bytes,
            file_name=f"{service_number}_BPSC_Report.xlsx" if service_number else "BPSC_Report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            type="secondary",
        )


def _to_cents(value: Any) -> int | None:
    """Convert a value to integer cents (2-decimal, half-up). Returns None if not numeric."""
    if value is None:
        return None
    # Handle pandas/numpy NaN
    try:
        if pd.isna(value):
            return None
    except Exception:
        pass
    if isinstance(value, str):
        text = value.strip().replace(",", "")
        if not text:
            return None
    else:
        text = str(value)
    try:
        dec = Decimal(text)
    except (InvalidOperation, ValueError):
        return None
    # Guard against Decimal('NaN') / Infinity
    try:
        if not dec.is_finite():
            return None
    except Exception:
        return None
    dec = dec.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
    try:
        if not dec.is_finite():
            return None
    except Exception:
        return None
    return int(dec * 100)


def _subset_solutions_meet_in_middle(values_cents: list[int], target_cents: int, max_solutions: int = 10) -> list[list[int]]:
    """Find up to `max_solutions` subsets whose sum equals `target_cents`.

    Returns a list of solutions; each solution is a list of indices into `values_cents`.
    Uses meet-in-the-middle; suitable for n up to ~40.
    """
    n = len(values_cents)
    if n == 0:
        return []

    mid = n // 2
    left_vals = values_cents[:mid]
    right_vals = values_cents[mid:]

    left: list[tuple[int, int]] = [(0, 0)]  # (sum, mask)
    for i, v in enumerate(left_vals):
        left += [(s + v, mask | (1 << i)) for (s, mask) in left]

    right: list[tuple[int, int]] = [(0, 0)]
    for i, v in enumerate(right_vals):
        right += [(s + v, mask | (1 << i)) for (s, mask) in right]

    right.sort(key=lambda x: x[0])
    right_sums = [s for (s, _) in right]

    solutions: list[list[int]] = []
    for left_sum, left_mask in left:
        needed = target_cents - left_sum
        lo = bisect.bisect_left(right_sums, needed)
        hi = bisect.bisect_right(right_sums, needed)
        if lo == hi:
            continue

        for _, right_mask in right[lo:hi]:
            if left_mask == 0 and right_mask == 0:
                continue

            idxs: list[int] = []
            for i in range(len(left_vals)):
                if left_mask & (1 << i):
                    idxs.append(i)
            for i in range(len(right_vals)):
                if right_mask & (1 << i):
                    idxs.append(mid + i)

            solutions.append(idxs)
            if len(solutions) >= max_solutions:
                return solutions

    return solutions


def _subset_solutions_dp_positive(values_cents: list[int], target_cents: int, max_solutions: int = 10) -> list[list[int]]:
    """DP subset-sum for non-negative ints; returns up to `max_solutions` solutions.

    This considers the full list of rows (not a row-count-limited window), but
    the number of solutions returned is capped to keep output practical.
    """
    if target_cents < 0:
        return []
    if target_cents == 0:
        # Avoid enumerating infinite-like combinations (especially with zeros).
        # Return the first zero-value row as a minimal non-empty solution, if any.
        for idx, v in enumerate(values_cents):
            if v == 0:
                return [[idx]]
        return []

    # dp[sum] -> list of tuples of indices that create that sum
    dp: dict[int, list[tuple[int, ...]]] = {0: [tuple()]}

    for idx, v in enumerate(values_cents):
        if v < 0:
            # Not supported by this DP variant.
            return []
        if v == 0:
            # Skip zeros to prevent combinatorial explosion; they do not help reach a positive target.
            continue
        if v > target_cents:
            continue

        current_sums = list(dp.keys())
        for s in current_sums:
            new_sum = s + v
            if new_sum > target_cents:
                continue
            existing_paths = dp.get(s, [])
            if not existing_paths:
                continue

            bucket = dp.setdefault(new_sum, [])
            if len(bucket) >= max_solutions:
                continue

            remaining = max_solutions - len(bucket)
            for path in existing_paths[:remaining]:
                bucket.append(path + (idx,))

        if len(dp.get(target_cents, [])) >= max_solutions:
            break

    return [list(p) for p in dp.get(target_cents, []) if p]


def _subset_value_patterns_backtrack(
    value_to_row_indexes: dict[int, list[int]],
    target_cents: int,
    max_patterns: int = 100,
) -> list[dict[int, int]]:
    """Find unique solutions as value-count patterns (treating duplicates as interchangeable).

    Returns list of patterns: {value_cents: count_used}.
    This avoids generating multiple equivalent combinations when the sheet has
    duplicate numeric values.
    """
    if target_cents <= 0:
        return []

    items: list[tuple[int, int]] = []  # (value_cents, available_count)
    for v, rows in value_to_row_indexes.items():
        if v <= 0:
            continue
        items.append((v, len(rows)))

    # Heuristic: large values first improves pruning.
    items.sort(key=lambda x: x[0], reverse=True)

    patterns: list[dict[int, int]] = []
    no_solution: set[tuple[int, int]] = set()

    def dfs(i: int, remaining: int, current: dict[int, int]) -> bool:
        if len(patterns) >= max_patterns:
            return True
        if remaining == 0:
            patterns.append(current.copy())
            return len(patterns) >= max_patterns
        if i >= len(items) or remaining < 0:
            return False
        state = (i, remaining)
        if state in no_solution:
            return False

        value, avail = items[i]
        max_take = min(avail, remaining // value)
        # Try higher counts first to reach target faster.
        for take in range(max_take, -1, -1):
            if take:
                current[value] = take
            else:
                current.pop(value, None)
            done = dfs(i + 1, remaining - take * value, current)
            if done:
                return True

        no_solution.add(state)
        return False

    dfs(0, target_cents, {})
    return patterns


def _subset_first_value_pattern_memo(
    value_to_row_indexes: dict[int, list[int]],
    target_cents: int,
) -> dict[int, int] | None:
    """Return a single exact solution as a value->count pattern.

    This mirrors the user's desired behavior: find *one* exact combination and
    stop, which avoids printing many equivalent possibilities when duplicates
    exist.
    """
    if target_cents <= 0:
        return None

    counts: dict[int, int] = {}
    for v, rows in value_to_row_indexes.items():
        if v <= 0:
            continue
        counts[v] = len(rows)

    if not counts:
        return None

    unique_vals = sorted(counts.keys(), reverse=True)
    # Ensure recursion depth is sufficient for large unique value sets.
    sys.setrecursionlimit(max(5000, len(unique_vals) + 200))

    memo: dict[tuple[int, int], dict[int, int] | None] = {}

    def solve(remaining: int, idx: int) -> dict[int, int] | None:
        if remaining == 0:
            return {}
        if remaining < 0 or idx >= len(unique_vals):
            return None

        state = (remaining, idx)
        if state in memo:
            return memo[state]

        val = unique_vals[idx]
        max_count = min(counts[val], remaining // val)

        for take in range(max_count, -1, -1):
            res = solve(remaining - take * val, idx + 1)
            if res is not None:
                out = dict(res)
                if take > 0:
                    out[val] = take
                memo[state] = out
                return out

        memo[state] = None
        return None

    return solve(target_cents, 0)


def _find_subset_solutions(values_cents: list[int], target_cents: int, max_solutions: int = 10) -> list[list[int]]:
    """Find subsets that sum to target.

    - For small n, meet-in-the-middle is fast.
    - For larger n (entire column), use DP for non-negative values.
    """
    n = len(values_cents)
    if n == 0:
        return []

    # Fast path for small inputs.
    if n <= 44:
        return _subset_solutions_meet_in_middle(values_cents, target_cents, max_solutions=max_solutions)

    # DP supports non-negative integers. Most billing/amount sheets are non-negative.
    if any(v < 0 for v in values_cents):
        return []
    return _subset_solutions_dp_positive(values_cents, target_cents, max_solutions=max_solutions)


def render_subset_calculator() -> None:
    st.title("Subset Calculator")
    st.caption("Upload an Excel sheet and find row combinations whose numeric values sum to the target.")

    # Allows us to programmatically reset the file uploader.
    if "subset_upload_nonce" not in st.session_state:
        st.session_state["subset_upload_nonce"] = 0

    # Refresh button: clears inputs/results without a browser refresh.
    if st.button("Refresh", key="subset_refresh", type="secondary"):
        st.session_state["subset_upload_nonce"] = int(st.session_state.get("subset_upload_nonce", 0)) + 1
        for k in list(st.session_state.keys()):
            if k.startswith("subset_") and k != "subset_upload_nonce":
                st.session_state.pop(k, None)
        st.rerun()

    uploaded = st.file_uploader(
        "Upload Excel file",
        type=["xlsx", "xls"],
        accept_multiple_files=False,
        key=f"subset_excel_upload_{st.session_state['subset_upload_nonce']}",
    )

    mode_col1, mode_col2 = st.columns([2, 1])
    with mode_col1:
        solution_mode = st.radio(
            "Search Mode",
            options=["Find only one solution", "Find up to N solutions"],
            index=0,
            horizontal=True,
            key="subset_solution_mode",
        )
    with mode_col2:
        max_solutions = st.number_input(
            "N",
            min_value=1,
            value=10,
            step=1,
            disabled=(solution_mode != "Find up to N solutions"),
            key="subset_max_solutions",
        )

    col_b, col_c = st.columns(2)
    with col_b:
        value_col_num = st.number_input(
            "Column number for numerical value (1-based)",
            min_value=1,
            value=2,
            step=1,
            key="subset_value_col",
        )
    with col_c:
        target_value = st.number_input(
            "Target numerical value",
            value=0.00,
            step=0.01,
            format="%.2f",
            key="subset_target",
        )

    if uploaded is None:
        st.info("Upload an Excel file to begin.")
        return

    # Reset stored results when inputs change
    try:
        file_sig = hashlib.md5(uploaded.getvalue()).hexdigest()
    except Exception:
        file_sig = str(getattr(uploaded, "name", "uploaded"))
    ctx_key = f"{file_sig}|col={int(value_col_num)}|target={float(target_value):.2f}"
    if st.session_state.get("subset_ctx_key") != ctx_key:
        st.session_state["subset_ctx_key"] = ctx_key
        st.session_state.pop("subset_last_patterns", None)
        st.session_state.pop("subset_last_message", None)
        st.session_state.pop("subset_excluded_rows", None)

    try:
        df = pd.read_excel(uploaded)
    except Exception as e:
        st.error(f"Failed to read Excel file: {e}")
        return

    # Permanent fix: convert Excel serial dates to dd/mm/yyyy (only for columns whose name contains 'date').
    if not df.empty:
        def _maybe_fix_excel_serial_date_column(col_name: str, series: pd.Series) -> pd.Series:
            if series.dtype.kind not in ("i", "u", "f"):
                return series

            # Only apply to likely date columns to avoid corrupting numeric data.
            name = str(col_name).strip().lower()
            if "date" not in name:
                return series

            non_null = series.dropna()
            if non_null.empty:
                return series

            sample = non_null.head(200)
            # Candidate Excel serial dates are typically ~ 20000..60000 for modern years.
            # Also ensure values are near-integers.
            near_int = (sample - sample.round()).abs() < 1e-9
            in_range = (sample >= 20000) & (sample <= 60000)
            if float((near_int & in_range).mean()) < 0.8:
                return series

            dt = pd.to_datetime(series, unit="D", origin="1899-12-30", errors="coerce")
            if dt.notna().sum() == 0:
                return series
            return dt.dt.strftime("%d/%m/%Y")

        df = df.copy()
        for c in df.columns:
            df[c] = _maybe_fix_excel_serial_date_column(c, df[c])

    if df.empty:
        st.error("Excel file has no rows.")
        return

    st.subheader("Preview")
    st.dataframe(df.head(20), use_container_width=True)

    num_cols = df.shape[1]
    value_idx = int(value_col_num) - 1
    if value_idx < 0 or value_idx >= num_cols:
        st.error(f"Invalid numeric column number. File has {num_cols} columns.")
        return

    target_cents = _to_cents(target_value)
    if target_cents is None:
        st.error("Target value must be numeric.")
        return

    value_col_name = df.columns[value_idx]

    # Build a mapping from numeric value -> row indexes (to treat duplicates as interchangeable)
    value_to_rows: dict[int, list[int]] = {}
    negatives_found = False
    for row_i, raw in enumerate(df.iloc[:, value_idx].tolist()):
        cents = _to_cents(raw)
        if cents is None:
            continue
        if cents < 0:
            negatives_found = True
        value_to_rows.setdefault(cents, []).append(row_i)

    if not value_to_rows:
        st.error(f"No numeric values found in column {value_col_num} ({value_col_name}).")
        return

    if negatives_found:
        st.warning("Negative values were found in the numeric column. The subset solver currently ignores negative/zero values for combination search.")

    numeric_row_count = sum(len(v) for v in value_to_rows.values())
    if numeric_row_count > 44:
        st.warning(
            f"Found {numeric_row_count} numeric rows. Searching the entire column can take longer, especially if many values can combine to the target."
        )

    button_label = "Find combination" if solution_mode == "Find only one solution" else "Find combinations"
    find_clicked = st.button(button_label, type="primary", key="subset_find")
    if find_clicked:
        with st.spinner("Searching combinations..."):
            if solution_mode == "Find only one solution":
                patterns: list[dict[int, int]] = []
                single = _subset_first_value_pattern_memo(value_to_rows, target_cents)
                if single:
                    patterns = [single]
            else:
                patterns = _subset_value_patterns_backtrack(value_to_rows, target_cents, max_patterns=int(max_solutions))

        if not patterns:
            st.session_state["subset_last_patterns"] = []
            st.session_state["subset_last_message"] = "No exact combination found for the target value."
        else:
            st.session_state["subset_last_patterns"] = patterns
            st.session_state["subset_last_message"] = None
            st.session_state.pop("subset_excluded_rows", None)

    patterns = st.session_state.get("subset_last_patterns")
    last_message = st.session_state.get("subset_last_message")
    if last_message:
        st.error(last_message)
        return
    if not patterns:
        return

    if solution_mode == "Find only one solution" and len(patterns) == 1:
        st.success(f"Exact combination found for {target_value:,.2f}.")
    else:
        st.success(f"Found {len(patterns)} unique combination pattern(s) matching {target_value:,.2f}.")

    def _format_money(cents: int) -> str:
        return f"{(Decimal(cents) / Decimal(100)):,.2f}"

    rendered: list[tuple[pd.DataFrame, pd.DataFrame]] = []
    for sol_idx, pattern in enumerate(patterns, start=1):
        # Choose representative rows for each value (first N rows for that value)
        selected_indexes: list[int] = []
        replaceable_values: set[int] = set()
        for value_cents, count_used in pattern.items():
            rows_for_value = sorted(value_to_rows.get(value_cents, []))
            if len(rows_for_value) > count_used:
                replaceable_values.add(value_cents)
            selected_indexes.extend(rows_for_value[:count_used])

        # Classify and order selected rows: irreplaceable first, then replaceable
        ordered_selected: list[tuple[int, int]] = []  # (category, row_index)
        for row_index in sorted(selected_indexes):
            row_value = _to_cents(df.iloc[row_index, value_idx])
            category = 1 if (row_value in replaceable_values) else 0
            ordered_selected.append((category, row_index))
        ordered_selected.sort(key=lambda x: (x[0], x[1]))
        selected_row_indexes = [ri for (_, ri) in ordered_selected]

        selected_df = df.iloc[selected_row_indexes].copy()

        # Alternatives: other rows with the same value that could substitute
        alt_rows: list[pd.DataFrame] = []
        for value_cents in sorted(replaceable_values, reverse=True):
            rows_all = sorted(value_to_rows.get(value_cents, []))
            used_count = pattern.get(value_cents, 0)
            used_set = set(rows_all[:used_count])
            others = [r for r in rows_all if r not in used_set]
            if not others:
                continue
            alt_df = df.iloc[others].copy()
            alt_df.insert(0, "Alternative_For_Value", _format_money(value_cents))
            alt_rows.append(alt_df)
        alternatives_df = pd.concat(alt_rows, ignore_index=True) if alt_rows else pd.DataFrame()

        sum_cents = sum(v * c for v, c in pattern.items())
        sum_value = float(Decimal(sum_cents) / Decimal(100))
        title = "Solution" if len(patterns) == 1 else f"Solution {sol_idx}"
        st.subheader(f"{title} (Sum: {sum_value:,.2f})")
        st.caption("Irreplaceable rows shown first; replaceable (duplicate-value) rows after.")
        st.dataframe(selected_df, use_container_width=True)

        # In single-solution mode, allow user to exclude selected rows and re-analyse for another possibility.
        if solution_mode == "Find only one solution" and len(patterns) == 1:
            excluded_set = set(st.session_state.get("subset_excluded_rows", []) or [])
            editor_df = selected_df.copy()
            editor_df.insert(0, "Exclude_Next_Search", editor_df.index.to_series().apply(lambda i: i in excluded_set))
            st.caption("Tick rows to exclude, then click 'Find another possibility'.")
            edited = st.data_editor(
                editor_df,
                use_container_width=True,
                key="subset_selected_editor",
                disabled=False,
            )
            try:
                new_excluded = edited.index[edited["Exclude_Next_Search"] == True].tolist()  # noqa: E712
            except Exception:
                new_excluded = []
            st.session_state["subset_excluded_rows"] = new_excluded

            reanalyse_clicked = st.button("Find another possibility", key="subset_reanalyse")
            if reanalyse_clicked:
                with st.spinner("Re-analysing with your selection..."):
                    filtered: dict[int, list[int]] = {}
                    excluded = set(st.session_state.get("subset_excluded_rows", []) or [])
                    for v, rows in value_to_rows.items():
                        kept = [r for r in rows if r not in excluded]
                        if kept:
                            filtered[v] = kept

                    current_pattern = patterns[0]

                    # First attempt: find a (possibly) new solution under exclusions.
                    candidate = _subset_first_value_pattern_memo(filtered, target_cents)
                    next_pattern: dict[int, int] | None = None
                    if candidate and candidate != current_pattern:
                        next_pattern = candidate
                    else:
                        # If exclusions are empty or candidate matches current, try to find the next distinct pattern.
                        try_patterns = _subset_value_patterns_backtrack(filtered, target_cents, max_patterns=25)
                        for p in try_patterns:
                            if p != current_pattern:
                                next_pattern = p
                                break

                    if next_pattern:
                        st.session_state["subset_last_patterns"] = [next_pattern]
                        st.session_state["subset_last_message"] = None
                        st.session_state.pop("subset_excluded_rows", None)
                        st.rerun()
                    else:
                        st.info("No other possible outcome based on your selection. Showing the same solution.")

            if not alternatives_df.empty:
                st.caption("Replaceable rows (alternatives for duplicate values used in the solution)")
                st.dataframe(alternatives_df, use_container_width=True)

            rendered.append((selected_df, alternatives_df))

    # Build downloadable Excel: one sheet per solution, with alternatives placed below selected rows
    output = io.BytesIO()
    try:
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            for sol_idx, (sel_df, alt_df) in enumerate(rendered, start=1):
                sheet_name = f"Solution_{sol_idx}"[:31]

                cols: list[str] = list(sel_df.columns)
                for c in list(alt_df.columns):
                    if c not in cols:
                        cols.append(c)

                sel_out = sel_df.reindex(columns=cols)
                if alt_df.empty:
                    sheet_df = sel_out
                else:
                    alt_out = alt_df.reindex(columns=cols)
                    blank_row = pd.DataFrame([{c: None for c in cols}])
                    label_row = pd.DataFrame([{cols[0]: "ALTERNATIVES (REPLACEABLE ROWS)"}], columns=cols)
                    sheet_df = pd.concat([sel_out, blank_row, label_row, alt_out], ignore_index=True)

                sheet_df.to_excel(writer, index=False, sheet_name=sheet_name)
        output.seek(0)
    except Exception as e:
        st.error(f"Failed to generate Excel output: {e}")
        return

    st.download_button(
        label="Download Excel Output",
        data=output.getvalue(),
        file_name="subset_calculator_output.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="secondary",
    )


# Sidebar navigation
selected_app = st.sidebar.radio(
    "Select Calculator",
    options=["Energy Adjustment Calculator", "BPSC Calculator", "Subset Calculator"],
    index=0,
)

# If user chooses BPSC, render that page and stop before executing the
# (large) Energy Adjustment UI below.
if selected_app == "BPSC Calculator":
    render_auto_update_sidebar()
    render_bpsc_calculator()
    render_footer("BPSC Calculator - Streamlit Version")
    st.stop()

if selected_app == "Subset Calculator":
    render_auto_update_sidebar()
    render_subset_calculator()
    render_footer("Subset Calculator - Streamlit Version")
    st.stop()

# Title and author
col1, col2 = st.columns([3, 1])
with col1:
    st.title("Energy Adjustment Calculator")
with col2:
    st.markdown(f"**Author: {PDF_AUTHOR_NAME}**")

# Initialize session state
if 'processed_data' not in st.session_state:
    st.session_state.processed_data = None
if 'error_message' not in st.session_state:
    st.session_state.error_message = None

def process_energy_data(generated_files, cpp_files, consumed_files,
                       enable_iex, enable_cpp, t_and_d_loss, cpp_t_and_d_loss,
                       consumer_number, consumer_name, multiplication_factor, tariff_selection,
                       auto_detect_month, month, year, date_filter):
    """Process energy data files and return calculation results"""
    try:
        uploaded_artifacts = []
        # Ensure file lists are not None
        if not enable_iex:
            generated_files = []
        if not enable_cpp:
            cpp_files = []
            
        # Process I.E.X generated energy Excel files (if provided)
        gen_df = None
        if generated_files and enable_iex and len(generated_files) > 0:
            gen_dfs = []
            for idx, gen_file in enumerate(generated_files, start=1):
                artifact = capture_uploaded_artifact(gen_file, 'iex', idx)
                temp_df = pd.read_excel(io.BytesIO(artifact['bytes']), header=0)
                if temp_df.shape[1] < 3:
                    return {'success': False, 'error': f"Generated energy Excel file '{gen_file.name}' must have at least 3 columns: Date, Time, and Energy in MW."}

                # Add filename to help with debugging
                temp_df['Source_File'] = artifact['original_name']
                gen_dfs.append(temp_df)
                uploaded_artifacts.append(artifact)
            
            # Combine all generated energy dataframes
            if gen_dfs:
                gen_df = pd.concat(gen_dfs, ignore_index=True)
                gen_df = gen_df.iloc[:, :3]
                gen_df.columns = ['Date', 'Time', 'Energy_MW']
                # Strip whitespace from Date and Time columns
                gen_df['Date'] = gen_df['Date'].astype(str).str.strip()
                gen_df['Time'] = gen_df['Time'].astype(str).str.strip()
                
                # Convert Energy_MW to numeric, handling string values
                gen_df['Energy_MW'] = pd.to_numeric(gen_df['Energy_MW'], errors='coerce')
                nan_count = gen_df['Energy_MW'].isna().sum()
                if nan_count > 0:
                    st.warning(f"{nan_count} non-numeric Energy_MW values found in I.E.X files and converted to NaN")
                
                # Standardize date format to yyyy-mm-dd for robust filtering
                gen_df['Date'] = pd.to_datetime(gen_df['Date'], errors='coerce', dayfirst=True)
                gen_df['Source_Type'] = 'I.E.X'

        # Process C.P.P (Captive Power Purchase) files (if provided)
        cpp_df = None
        if cpp_files and enable_cpp and len(cpp_files) > 0:
            cpp_dfs = []
            for idx, cpp_file in enumerate(cpp_files, start=1):
                artifact = capture_uploaded_artifact(cpp_file, 'cpp', idx)
                temp_df = pd.read_excel(io.BytesIO(artifact['bytes']), header=0)
                if temp_df.shape[1] < 3:
                    return {'success': False, 'error': f"C.P.P energy Excel file '{cpp_file.name}' must have at least 3 columns: Date, Time, and Energy in MW."}
                
                # Add filename to help with debugging
                temp_df['Source_File'] = artifact['original_name']
                temp_df['Source_Type'] = 'C.P.P'
                cpp_dfs.append(temp_df)
                uploaded_artifacts.append(artifact)
            
            # Process C.P.P data if files were uploaded
            if cpp_dfs:
                cpp_df = pd.concat(cpp_dfs, ignore_index=True)
                cpp_df = cpp_df.iloc[:, :3]
                cpp_df.columns = ['Date', 'Time', 'Energy_MW']
                cpp_df['Date'] = cpp_df['Date'].astype(str).str.strip()
                cpp_df['Time'] = cpp_df['Time'].astype(str).str.strip()
                
                # Convert Energy_MW to numeric, handling string values
                cpp_df['Energy_MW'] = pd.to_numeric(cpp_df['Energy_MW'], errors='coerce')
                nan_count = cpp_df['Energy_MW'].isna().sum()
                if nan_count > 0:
                    st.warning(f"{nan_count} non-numeric Energy_MW values found in C.P.P files and converted to NaN")
                
                cpp_df['Date'] = pd.to_datetime(cpp_df['Date'], errors='coerce', dayfirst=True)
                cpp_df['Source_Type'] = 'C.P.P'

        # Combine I.E.X and C.P.P data if both exist
        if gen_df is not None and cpp_df is not None:
            combined_gen_df = pd.concat([gen_df, cpp_df], ignore_index=True)
        elif gen_df is not None:
            combined_gen_df = gen_df
        elif cpp_df is not None:
            combined_gen_df = cpp_df
        else:
            return {'success': False, 'error': "No valid generation energy files were found."}
        
        gen_df = combined_gen_df

        # Function to convert month number to month name
        def get_month_name(month_num):
            month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                          'July', 'August', 'September', 'October', 'November', 'December']
            try:
                return month_names[int(month_num) - 1]
            except (ValueError, IndexError):
                return str(month_num)

        # Auto-detect month and year if enabled
        auto_detect_info = ""
        if auto_detect_month and not (month and year):
            # Extract unique months and years from the data
            unique_months = gen_df['Date'].dt.month.unique()
            unique_years = gen_df['Date'].dt.year.unique()
            
            if len(unique_months) == 1 and not month:
                month = str(int(unique_months[0]))
                st.info(f"Auto-detected month: {month} ({get_month_name(month)})")
            elif len(unique_months) > 1 and not month:
                # If multiple months, use the most frequent one
                month = str(int(gen_df['Date'].dt.month.value_counts().idxmax()))
                st.info(f"Multiple months detected, using most frequent: {month} ({get_month_name(month)})")
            
            if len(unique_years) == 1 and not year:
                year = str(int(unique_years[0]))
                st.info(f"Auto-detected year: {year}")
            elif len(unique_years) > 1 and not year:
                # If multiple years, use the most frequent one
                year = str(int(gen_df['Date'].dt.year.value_counts().idxmax()))
                st.info(f"Multiple years detected, using most frequent: {year}")
                
            # Add information to be displayed in PDF
            cpp_count = len(cpp_files) if cpp_files else 0
            iex_count = len(generated_files) if generated_files else 0
            auto_detect_info = f"Auto-detected from {iex_count} I.E.X, {cpp_count} C.P.P, and {len(consumed_files)} consumed files"

        # Process multiple consumed energy Excel files
        cons_dfs = []
        for idx, cons_file in enumerate(consumed_files, start=1):
            artifact = capture_uploaded_artifact(cons_file, 'consumption', idx)
            temp_df = pd.read_excel(io.BytesIO(artifact['bytes']), header=0)
            if temp_df.shape[1] < 3:
                return {'success': False, 'error': f"Consumed energy Excel file '{cons_file.name}' must have at least 3 columns: Date, Time, and Energy in kWh."}
            
            # Add filename to help with debugging
            temp_df['Source_File'] = artifact['original_name']
            cons_dfs.append(temp_df)
            uploaded_artifacts.append(artifact)
        
        # Combine all consumed energy dataframes
        if not cons_dfs:
            return {'success': False, 'error': "No valid consumed energy Excel files were found."}
        
        cons_df = pd.concat(cons_dfs, ignore_index=True)
        cons_df = cons_df.iloc[:, :3]
        cons_df.columns = ['Date', 'Time', 'Energy_kWh']
        # Strip whitespace from Date and Time columns
        cons_df['Date'] = cons_df['Date'].astype(str).str.strip()
        cons_df['Time'] = cons_df['Time'].astype(str).str.strip()
        # Standardize date format to yyyy-mm-dd for robust filtering
        cons_df['Date'] = pd.to_datetime(cons_df['Date'], errors='coerce', dayfirst=True)
        
        # Apply date filtering logic (simplified version)
        filtered_gen = gen_df.copy()
        filtered_cons = cons_df.copy()
        
        if year and month:
            try:
                year_int = int(float(year))
                month_int = int(float(month))
                
                # Filter generation data
                filtered_gen = filtered_gen[
                    (filtered_gen['Date'].dt.year == year_int) & 
                    (filtered_gen['Date'].dt.month == month_int)
                ]
                
                # Filter consumption data
                filtered_cons = filtered_cons[
                    (filtered_cons['Date'].dt.year == year_int) & 
                    (filtered_cons['Date'].dt.month == month_int)
                ]
                
            except ValueError:
                return {'success': False, 'error': f"Invalid year or month value. Year: '{year}', Month: '{month}'"}
        
        if date_filter:
            try:
                date_obj = pd.to_datetime(date_filter, dayfirst=True)
                filtered_gen = filtered_gen[filtered_gen['Date'] == date_obj]
                filtered_cons = filtered_cons[filtered_cons['Date'] == date_obj]
            except Exception:
                return {'success': False, 'error': f"Invalid date format for filter: {date_filter}. Use dd/mm/yyyy."}
        
        if (year or month or date_filter) and (filtered_gen.empty or filtered_cons.empty):
            return {'success': False, 'error': "No data found for the selected filters."}
        
        gen_df = filtered_gen
        cons_df = filtered_cons
        
        # Convert MW to kWh (MW * 250 for 15-minute intervals)
        gen_df['Energy_kWh'] = gen_df['Energy_MW'] * 250
        
        # Apply T&D losses based on source type
        def apply_td_loss(row):
            if row['Source_Type'] == 'I.E.X':
                return row['Energy_kWh'] * (1 - t_and_d_loss / 100) if t_and_d_loss > 0 else row['Energy_kWh']
            elif row['Source_Type'] == 'C.P.P':
                return row['Energy_kWh'] * (1 - cpp_t_and_d_loss / 100) if cpp_t_and_d_loss > 0 else row['Energy_kWh']
            else:
                return row['Energy_kWh']
        
        gen_df['After_Loss'] = gen_df.apply(apply_td_loss, axis=1)
        
        # Create slot time and date columns
        def slot_time_range(row):
            t = str(row['Time']).strip()
            if '-' in t:
                return t
            try:
                start = pd.to_datetime(t, format='%H:%M').time()
                end_dt = (pd.Timestamp.combine(pd.Timestamp.today(), start) + pd.Timedelta(minutes=15)).time()
                return f"{start.strftime('%H:%M')} - {end_dt.strftime('%H:%M')}"
            except Exception:
                return t
        
        gen_df['Slot_Time'] = gen_df.apply(slot_time_range, axis=1)
        gen_df['Slot_Time'] = gen_df['Slot_Time'].replace({'23:45 - 24:00': '23:45 - 00:00'})
        gen_df['Slot_Date'] = gen_df['Date'].dt.strftime('%d/%m/%Y')
        
        # Apply same processing to consumption data
        cons_df['Energy_kWh'] = pd.to_numeric(cons_df['Energy_kWh'], errors='coerce') * multiplication_factor
        cons_df['Slot_Time'] = cons_df.apply(slot_time_range, axis=1)
        cons_df['Slot_Time'] = cons_df['Slot_Time'].replace({'23:45 - 24:00': '23:45 - 00:00'})
        cons_df['Slot_Date'] = cons_df['Date'].dt.strftime('%d/%m/%Y')

        # Sequential adjustment logic: First I.E.X, then C.P.P
        # Separate I.E.X and C.P.P data for sequential adjustment
        iex_df = gen_df[gen_df['Source_Type'] == 'I.E.X'].copy() if enable_iex else pd.DataFrame()
        cpp_df_only = gen_df[gen_df['Source_Type'] == 'C.P.P'].copy() if enable_cpp else pd.DataFrame()
        
        # Build all slot combinations from consumption and both generation sources
        cons_slots_set = set((d, t) for d, t in zip(cons_df['Slot_Date'], cons_df['Slot_Time']))
        iex_slots_set = set((d, t) for d, t in zip(iex_df['Slot_Date'], iex_df['Slot_Time'])) if not iex_df.empty else set()
        cpp_slots_set = set((d, t) for d, t in zip(cpp_df_only['Slot_Date'], cpp_df_only['Slot_Time'])) if not cpp_df_only.empty else set()
        
        all_slots = pd.DataFrame(
            list(cons_slots_set | iex_slots_set | cpp_slots_set),
            columns=['Slot_Date', 'Slot_Time'])
        
        # Merge consumption data first
        merged = pd.merge(all_slots, cons_df[['Slot_Date', 'Slot_Time', 'Energy_kWh']], on=['Slot_Date', 'Slot_Time'], how='left')
        merged['Energy_kWh_cons'] = merged['Energy_kWh'].fillna(0)
        merged.drop('Energy_kWh', axis=1, inplace=True)
        
        # Merge I.E.X data
        if not iex_df.empty:
            iex_merge = iex_df[['Slot_Date', 'Slot_Time', 'After_Loss', 'Energy_kWh']].copy()
            iex_merge.columns = ['Slot_Date', 'Slot_Time', 'IEX_After_Loss', 'IEX_Energy_kWh']
            merged = pd.merge(merged, iex_merge, on=['Slot_Date', 'Slot_Time'], how='left')
            merged['IEX_After_Loss'] = merged['IEX_After_Loss'].fillna(0)
            merged['IEX_Energy_kWh'] = merged['IEX_Energy_kWh'].fillna(0)
        else:
            merged['IEX_After_Loss'] = 0
            merged['IEX_Energy_kWh'] = 0
        
        # Merge C.P.P data
        if not cpp_df_only.empty:
            cpp_merge = cpp_df_only[['Slot_Date', 'Slot_Time', 'After_Loss', 'Energy_kWh']].copy()
            cpp_merge.columns = ['Slot_Date', 'Slot_Time', 'CPP_After_Loss', 'CPP_Energy_kWh']
            merged = pd.merge(merged, cpp_merge, on=['Slot_Date', 'Slot_Time'], how='left')
            merged['CPP_After_Loss'] = merged['CPP_After_Loss'].fillna(0)
            merged['CPP_Energy_kWh'] = merged['CPP_Energy_kWh'].fillna(0)
        else:
            merged['CPP_After_Loss'] = 0
            merged['CPP_Energy_kWh'] = 0
        
        # Sequential Adjustment Calculation
        # Step 1: I.E.X adjustment first
        merged['IEX_Adjustment'] = merged[['IEX_After_Loss', 'Energy_kWh_cons']].min(axis=1)
        merged['IEX_Excess'] = (merged['IEX_After_Loss'] - merged['Energy_kWh_cons']).apply(lambda x: max(0, x))
        
        # Step 2: Calculate remaining consumption after I.E.X adjustment
        merged['Remaining_Consumption'] = (merged['Energy_kWh_cons'] - merged['IEX_Adjustment']).apply(lambda x: max(0, x))
        
        # Step 3: C.P.P adjustment with remaining consumption
        merged['CPP_Adjustment'] = merged[['CPP_After_Loss', 'Remaining_Consumption']].min(axis=1)
        merged['CPP_Excess'] = (merged['CPP_After_Loss'] - merged['Remaining_Consumption']).apply(lambda x: max(0, x))
        
        # Step 4: Total calculations
        merged['Total_Excess'] = merged['IEX_Excess'] + merged['CPP_Excess']
        merged['Total_Generated_After_Loss'] = merged['IEX_After_Loss'] + merged['CPP_After_Loss']
        merged['Total_Generated_Before_Loss'] = merged['IEX_Energy_kWh'] + merged['CPP_Energy_kWh']
        
        # For backward compatibility with existing PDF code
        merged['After_Loss'] = merged['Total_Generated_After_Loss']
        merged['Energy_kWh_gen'] = merged['Total_Generated_Before_Loss']
        merged['Excess'] = merged['Total_Excess']
        
        # Track missing slots for reporting
        merged['Missing_Info'] = ''
        merged['is_missing_iex'] = ~merged.apply(lambda row: (row['Slot_Date'], row['Slot_Time']) in iex_slots_set, axis=1)
        merged['is_missing_cpp'] = ~merged.apply(lambda row: (row['Slot_Date'], row['Slot_Time']) in cpp_slots_set, axis=1)
        merged['is_missing_cons'] = ~merged.apply(lambda row: (row['Slot_Date'], row['Slot_Time']) in cons_slots_set, axis=1)
        
        if enable_iex:
            merged.loc[merged['is_missing_iex'], 'Missing_Info'] = (
                merged.loc[merged['is_missing_iex'], 'Missing_Info'].astype(str) + '[Missing in I.E.X] '
            )
        if enable_cpp:
            merged.loc[merged['is_missing_cpp'], 'Missing_Info'] = (
                merged.loc[merged['is_missing_cpp'], 'Missing_Info'].astype(str) + '[Missing in C.P.P] '
            )
        merged.loc[merged['is_missing_cons'], 'Missing_Info'] = (
            merged.loc[merged['is_missing_cons'], 'Missing_Info'].astype(str) + '[Missing in CONSUMED] '
        )
        merged.drop(['is_missing_iex', 'is_missing_cpp', 'is_missing_cons'], axis=1, inplace=True)
        
        # Sort merged data chronologically by Slot_Date and Slot_Time
        def slot_time_to_minutes(slot_time):
            try:
                start = slot_time.split('-')[0].strip()
                h, m = map(int, start.split(':'))
                return h * 60 + m
            except Exception:
                return 0
        
        # Be flexible with date parsing: accept various formats and day-first entries
        merged['Slot_Date_dt'] = _robust_to_datetime(merged['Slot_Date'])
        merged['Slot_Time_min'] = merged['Slot_Time'].apply(slot_time_to_minutes)
        merged = merged.sort_values(['Slot_Date_dt', 'Slot_Time_min']).reset_index(drop=True)
        
        # Add TOD (Time of Day) classification
        def classify_tod(slot_time):
            try:
                start_time = slot_time.split('-')[0].strip()
                hour, minute = map(int, start_time.split(':'))
                
                # Morning peak: 6:00 AM - 10:00 AM (C1)
                if 6 <= hour < 10:
                    return 'C1'
                # Evening peak: 6:00 PM - 10:00 PM (C2)
                elif 18 <= hour < 22:
                    return 'C2'
                # Normal hours: 5:00 AM - 6:00 AM + 10:00 AM to 6:00 PM (C4)
                elif (5 <= hour < 6) or (10 <= hour < 18):
                    return 'C4'
                # Night hours: 22:00 PM to 5:00 AM (C5)
                elif (hour >= 22) or (hour < 5):
                    return 'C5'
                else:
                    return 'Unknown'
            except Exception:
                return 'Unknown'
        
        merged['TOD_Category'] = merged['Slot_Time'].apply(classify_tod)
        
        # Clean up temporary columns
        merged.drop(['Slot_Date_dt', 'Slot_Time_min'], axis=1, inplace=True)
        
        # Calculate totals
        sum_injection = merged['Energy_kWh_gen'].sum()
        total_generated_after_loss = merged['After_Loss'].sum()
        total_consumed = merged['Energy_kWh_cons'].sum()
        total_excess = merged['Total_Excess'].sum()
        
        # For PDF, show all slots or only excess slots
        merged_excess = merged[merged['Total_Excess'] > 0].copy()
        merged_all = merged.copy()
        
        # Count unique days
        unique_days_gen = merged['Slot_Date'].nunique()
        unique_days_cons = merged['Slot_Date'].nunique()
        
        excess_status = 'Excess' if total_excess > 0 else 'No Excess'
        
        # Calculate TOD-wise excess for financial calculations
        tod_excess = merged.groupby('TOD_Category')['Total_Excess'].sum().reset_index()
        
        # Helper function for consistent rounding throughout the application
        def round_kwh_financial(value):
            return int(value + 0.5) if value >= 0 else int(value - 0.5)
        
        # Round the total for financial calculations to match table display values
        total_excess_financial_rounded = round_kwh_financial(total_excess)
        
        # Resolve tariff-specific rates using selected period
        tariff_rates = resolve_tariff_rates(tariff_selection, month, year)
        base_rate = tariff_rates['base_rate']
        c1_c2_rate = tariff_rates['c1_c2_rate']
        c5_rate = tariff_rates['c5_rate']
        cross_subsidy_rate = tariff_rates['cross_subsidy_rate']
        wheeling_unit_rate = tariff_rates['wheeling_rate']

        # Base rate for all excess energy using rounded values
        base_amount = total_excess_financial_rounded * base_rate
        
        # Additional charges for specific TOD categories using rounded values
        c1_c2_excess_raw = tod_excess.loc[tod_excess['TOD_Category'].isin(['C1', 'C2']), 'Total_Excess'].sum()
        c1_c2_excess = round_kwh_financial(c1_c2_excess_raw)
        c1_c2_additional = c1_c2_excess * c1_c2_rate  # rupees per kWh
        
        c5_excess_raw = tod_excess.loc[tod_excess['TOD_Category'] == 'C5', 'Total_Excess'].sum()
        c5_excess = round_kwh_financial(c5_excess_raw)
        c5_additional = c5_excess * c5_rate  # rupees per kWh
        
        # Calculate total amount
        total_amount = base_amount + c1_c2_additional + c5_additional

        # Calculate E-Tax (5%)
        etax = total_amount * 0.05

        # Calculate total amount with E-Tax
        total_with_etax = total_amount + etax

        # Calculate IEX excess for specific charges using rounded values
        iex_excess_financial_raw = merged['IEX_Excess'].sum()
        iex_excess_financial = round_kwh_financial(iex_excess_financial_raw)

        # Calculate negative factors using rounded values
        etax_on_iex = total_excess_financial_rounded * 0.1
        cross_subsidy_surcharge = iex_excess_financial * cross_subsidy_rate  # Only for IEX excess

        additional_surcharge, additional_surcharge_breakdown, additional_surcharge_rate, additional_surcharge_period_label, additional_surcharge_note = calculate_monthly_additional_surcharge(
            month,
            year,
            iex_excess_financial_raw,
            iex_excess_financial,
        )
        
        # Wheeling charges based on revised reference addition and 2.34% deduction
        wheeling_components = compute_wheeling_components(
            total_excess_financial_rounded,
            t_and_d_loss,
        )
        wheeling_reference_kwh = wheeling_components['reference_kwh']
        wheeling_combined_kwh = wheeling_components['combined_kwh']
        wheeling_reduction_kwh = wheeling_components['reduction_kwh']
        wheeling_adjusted_kwh = wheeling_components['adjusted_kwh']
        wheeling_charges = wheeling_adjusted_kwh * wheeling_unit_rate
        
        # Calculate final amount to be collected (Additional Surcharge is brought in less like E-Tax on IEX and Cross Subsidy)
        final_amount = total_with_etax - (etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge)

        # Round up final amount to next highest value
        final_amount_rounded = math.ceil(final_amount)

        return {'success': True, 'data': {
            'merged_all': merged_all,
            'merged_excess': merged_excess,
            'sum_injection': sum_injection,
            'total_generated_after_loss': total_generated_after_loss,
            'total_consumed': total_consumed,
            'total_excess': total_excess,
            'excess_status': excess_status,
            'unique_days_gen': unique_days_gen,
            'unique_days_cons': unique_days_cons,
            'month': month,
            'year': year,
            'auto_detect_info': auto_detect_info,
            'enable_iex': enable_iex,
            'enable_cpp': enable_cpp,
            't_and_d_loss': t_and_d_loss,
            'cpp_t_and_d_loss': cpp_t_and_d_loss,
            'consumer_number': consumer_number,
            'consumer_name': consumer_name,
            'multiplication_factor': multiplication_factor,
            'tariff_selection': tariff_rates['tariff'],
            'tariff_window_label': tariff_rates['window_label'],
            'tariff_rates': tariff_rates,
            # Financial calculation results
            'total_excess_financial_rounded': total_excess_financial_rounded,
            'base_rate': base_rate,
            'tariff_c1_c2_rate': c1_c2_rate,
            'tariff_c5_rate': c5_rate,
            'tariff_cross_subsidy_rate': cross_subsidy_rate,
            'tariff_wheeling_rate': wheeling_unit_rate,
            'base_amount': base_amount,
            'c1_c2_excess': c1_c2_excess,
            'c1_c2_additional': c1_c2_additional,
            'c5_excess': c5_excess,
            'c5_additional': c5_additional,
            'total_amount': total_amount,
            'etax': etax,
            'total_with_etax': total_with_etax,
            'iex_excess_financial': iex_excess_financial,
            'etax_on_iex': etax_on_iex,
            'cross_subsidy_surcharge': cross_subsidy_surcharge,
            'additional_surcharge': additional_surcharge,
            'additional_surcharge_breakdown': additional_surcharge_breakdown,
            'additional_surcharge_rate': additional_surcharge_rate,
            'additional_surcharge_period_label': additional_surcharge_period_label,
            'additional_surcharge_note': additional_surcharge_note,
            'additional_surcharge_kwh': iex_excess_financial,
            'additional_surcharge_kwh_raw': iex_excess_financial_raw,
            'wheeling_charges': wheeling_charges,
            'wheeling_reference_kwh': wheeling_reference_kwh,
            'wheeling_combined_kwh': wheeling_combined_kwh,
            'wheeling_reduction_kwh': wheeling_reduction_kwh,
            'wheeling_adjusted_kwh': wheeling_adjusted_kwh,
            'final_amount': final_amount,
            'final_amount_rounded': final_amount_rounded,
            'message': "Data processing completed successfully!",
            'uploaded_artifacts': uploaded_artifacts,
        }}
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def generate_custom_filename(base_name, consumer_number, consumer_name, month=None, year=None, extension=".pdf"):
    """Generate custom filename with optional extension and suffix logic."""
    extension = extension if str(extension).startswith('.') else f".{extension}"
    # Get last 3 digits of service number
    last_3_digits = str(consumer_number)[-3:] if len(str(consumer_number)) >= 3 else str(consumer_number)

    # Clean consumer name for filename (remove special characters)
    clean_name = "".join(c for c in str(consumer_name) if c.isalnum() or c in (' ', '-', '_')).strip()
    clean_name = clean_name.replace(' ', '_') or "consumer"

    # Format month and year for filename (short date format)
    date_suffix = ""
    if month and year:
        try:
            month_num = int(float(month))
            year_num = int(float(year))
            # Format as MM_YY (e.g., 07_25 for July 2025)
            date_suffix = f"_{month_num:02d}_{year_num % 100:02d}"
        except Exception:
            date_suffix = ""  # If conversion fails, skip date suffix

    base_filename = f"{last_3_digits}_{clean_name}{date_suffix}"

    suffix = ""
    name_hint = (base_name or "").lower()
    if 'excess_only' in name_hint:
        suffix = "_excess_only"
    elif 'all_slots' in name_hint:
        suffix = "_all_slots"
    elif 'daywise' in name_hint:
        suffix = "_daywise"
    elif name_hint:
        safe_suffix = "".join(c for c in name_hint if c.isalnum() or c in ('_', '-')).strip('_')
        if safe_suffix:
            suffix = f"_{safe_suffix}"

    return f"{base_filename}{suffix}{extension}"


def append_hash_suffix(filename, hash_hex, label="sha256"):
    """Append a short tamper-evident hash tag to a filename."""
    if not hash_hex:
        return filename
    tag = hash_hex[:12]
    if not tag:
        return filename
    path = Path(filename)
    stem = path.stem or filename
    suffix = path.suffix
    return f"{stem}_{label}_{tag}{suffix}"


def capture_uploaded_artifact(uploaded_file, category, index):
    """Capture bytes and hash metadata for an uploaded Excel file."""
    file_bytes = uploaded_file.getvalue()
    uploaded_file.seek(0)
    sha256_hash = hashlib.sha256(file_bytes).hexdigest()
    extension = Path(uploaded_file.name).suffix or ".xlsx"
    safe_category = "".join(c for c in str(category) if c.isalnum() or c in ('_', '-')).strip('_') or "source"
    base_name = f"{safe_category}_upload_{index}"
    return {
        'category': safe_category,
        'sequence': index,
        'original_name': uploaded_file.name,
        'bytes': file_bytes,
        'hash': sha256_hash,
        'extension': extension,
        'base_name': base_name,
    }


def build_consumer_slug(consumer_number, consumer_name):
    """Create a reusable slug for bundle filenames."""
    last_3_digits = str(consumer_number)[-3:] if len(str(consumer_number)) >= 3 else str(consumer_number)
    clean_name = "".join(c for c in str(consumer_name) if c.isalnum() or c in (' ', '-', '_')).strip()
    clean_name = clean_name.replace(' ', '_') or "consumer"
    return f"{last_3_digits}_{clean_name}"

def generate_detailed_pdf(data, pdf_data, pdf_type):
    """Generate detailed PDF with complete table data and calculations"""
    try:
        # Import datetime for timestamp
        from datetime import datetime

        # Ensure additional surcharge defaults exist at function scope to avoid unbound references
        additional_surcharge = 0.0
        additional_surcharge_breakdown = []

        # Resolve tariff details for descriptive sections and fallback calculations
        tariff_selection = data.get('tariff_selection', TARIFF_OPTIONS[0])
        tariff_rates = data.get('tariff_rates') or resolve_tariff_rates(
            tariff_selection,
            data.get('month'),
            data.get('year'),
        )
        base_rate = data.get('base_rate', tariff_rates['base_rate'])
        c1_c2_rate = data.get('tariff_c1_c2_rate', tariff_rates['c1_c2_rate'])
        c5_rate = data.get('tariff_c5_rate', tariff_rates['c5_rate'])
        cross_subsidy_rate = data.get('tariff_cross_subsidy_rate', tariff_rates['cross_subsidy_rate'])
        wheeling_rate = data.get('tariff_wheeling_rate', tariff_rates['wheeling_rate'])
        
        pdf = AuthorPDF()
        pdf.set_margins(20, 20, 20)  # Set proper margins: left, top, right (20mm each)
        pdf.set_auto_page_break(auto=True, margin=20)  # Auto page break with bottom margin
        pdf.add_page()
        
        # FIRST PAGE - DESCRIPTION AND INFORMATION ONLY
        pdf.set_font('Arial', 'B', 16)  # Larger title font
        title = f"Energy Adjustment {pdf_type.replace('_', ' ').title()} Report"
        pdf.cell(0, 15, title, ln=True, align='C')
        pdf.ln(10)
        
        # Consumer Information Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Consumer Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f"Consumer Number: {data['consumer_number']}", ln=True)
        pdf.cell(0, 8, f"Consumer Name: {data['consumer_name']}", ln=True)
        pdf.cell(0, 8, f"Multiplication Factor (Consumed Energy): {data.get('multiplication_factor', 1)}", ln=True)
        pdf.cell(0, 8, f"Tariff: {tariff_selection} ({tariff_rates.get('window_label', 'Latest')})", ln=True)
        pdf.ln(5)
        
        # Technical Parameters Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Technical Parameters:', ln=True)
        pdf.set_font('Arial', '', 12)
        
        # Display T&D losses based on what sources are used
        if data.get('enable_iex') and data.get('enable_cpp'):
            pdf.cell(0, 8, f"I.E.X T&D Loss (%): {data.get('t_and_d_loss', 0)}", ln=True)
            pdf.cell(0, 8, f"C.P.P T&D Loss (%): {data.get('cpp_t_and_d_loss', 0)}", ln=True)
        elif data.get('enable_iex'):
            pdf.cell(0, 8, f"I.E.X T&D Loss (%): {data.get('t_and_d_loss', 0)}", ln=True)
        elif data.get('enable_cpp'):
            pdf.cell(0, 8, f"C.P.P T&D Loss (%): {data.get('cpp_t_and_d_loss', 0)}", ln=True)
        pdf.ln(5)
        
        # Data Sources Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Data Sources:', ln=True)
        pdf.set_font('Arial', '', 12)
        # Add count of files used (placeholder - can be enhanced with actual file count)
        if data.get('enable_iex'):
            pdf.cell(0, 8, f'Generated Energy Files (I.E.X): Available', ln=True)
        if data.get('enable_cpp'):
            pdf.cell(0, 8, f'Generated Energy Files (C.P.P): Available', ln=True)
        pdf.cell(0, 8, f'Consumed Energy Files: Available', ln=True)
        pdf.ln(5)
        
        # Report Information Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Report Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f"Report Generated: {datetime.now().strftime('%d/%m/%Y at %H:%M:%S')}", ln=True)
        if data.get('month') and data.get('year'):
            pdf.cell(0, 8, f"Report Period: {data['month']}/{data['year']}", ln=True)
        if data.get('auto_detect_info'):
            pdf.cell(0, 8, f"Period Details: {data['auto_detect_info']}", ln=True)
        pdf.ln(10)
        
        # Calculation Methodology Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Calculation Methodology:', ln=True)
        pdf.set_font('Arial', '', 12)
        
        methodology_text = [
            "1. Energy Adjustment = Net Generated Energy - Adjusted Consumed Energy",
            "2. Adjusted Consumed Energy = Consumed Energy × Multiplication Factor",
            "3. Net Generated Energy = Generated Energy × (1 - T&D Loss %)",
            "4. Financial calculations use Time of Day (TOD) rates",
            "5. All calculations maintain precision and consistency across tables and summaries"
        ]
        
        for line in methodology_text:
            pdf.cell(0, 6, line, ln=True)
        
        pdf.ln(10)
        
        # Add page break before tables
        pdf.add_page()
        
        # SECOND PAGE AND BEYOND - TABLES START HERE
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Energy Adjustment Tables', ln=True, align='C')
        pdf.ln(10)
        
        # Helper functions for safe formatting
        def safe_date_str(date_val):
            """Safely convert date to string"""
            if pd.isna(date_val):
                return ""
            return str(date_val)
        
        def format_time(time_val):
            """Format time value for display"""
            if pd.isna(time_val):
                return ""
            return str(time_val)
        
        # Function to add table headers with proper text wrapping
        def add_table_headers():
            is_dual_source = data.get('enable_iex') and data.get('enable_cpp')
            
            if is_dual_source:
                # Sequential adjustment table with detailed columns - improved headers
                pdf.set_font('Arial', 'B', 8)  # Multi-source headers: Font size 8
                
                # First row - main headers
                pdf.cell(16, 8, 'Date', 1, 0, 'C')
                pdf.cell(20, 8, 'Time', 1, 0, 'C')
                pdf.cell(12, 8, 'TOD', 1, 0, 'C')
                pdf.cell(18, 8, 'Consumed', 1, 0, 'C')
                pdf.cell(18, 8, 'IEX After', 1, 0, 'C')
                pdf.cell(16, 8, 'IEX', 1, 0, 'C')
                pdf.cell(18, 8, 'CPP After', 1, 0, 'C')
                pdf.cell(16, 8, 'CPP', 1, 0, 'C')
                pdf.cell(18, 8, 'Total', 1, 0, 'C')
                pdf.cell(12, 8, 'Missing', 1, 0, 'C')
                pdf.ln()
                
                # Second row - specifications
                pdf.cell(16, 8, '', 1, 0, 'C')
                pdf.cell(20, 8, '', 1, 0, 'C')
                pdf.cell(12, 8, '', 1, 0, 'C')
                pdf.cell(18, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(18, 8, 'Loss (kWh)', 1, 0, 'C')
                pdf.cell(16, 8, 'Excess', 1, 0, 'C')
                pdf.cell(18, 8, 'Loss (kWh)', 1, 0, 'C')
                pdf.cell(16, 8, 'Excess', 1, 0, 'C')
                pdf.cell(18, 8, 'Excess', 1, 0, 'C')
                pdf.cell(12, 8, 'Info', 1, 0, 'C')
                pdf.ln()
                
                # Third row - units
                pdf.cell(16, 8, '', 1, 0, 'C')
                pdf.cell(20, 8, '', 1, 0, 'C')
                pdf.cell(12, 8, '', 1, 0, 'C')
                pdf.cell(18, 8, '', 1, 0, 'C')
                pdf.cell(18, 8, '', 1, 0, 'C')
                pdf.cell(16, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(18, 8, '', 1, 0, 'C')
                pdf.cell(16, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(18, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(12, 8, '', 1, 0, 'C')
                pdf.ln()
                
                # Reset font to table data font after headers
                pdf.set_font('Arial', '', 9)  # Multi-source table content font size
                
            else:
                # Standard table for single source with properly formatted headers
                pdf.set_font('Arial', 'B', 9)  # Single-source headers: Font size 9
                
                # First row - main headers with borders
                pdf.cell(20, 8, 'Date', 1, 0, 'C')
                pdf.cell(25, 8, 'Time', 1, 0, 'C')  
                pdf.cell(15, 8, 'TOD', 1, 0, 'C')  
                pdf.cell(25, 8, 'Generated', 1, 0, 'C')
                pdf.cell(25, 8, 'Consumed', 1, 0, 'C') 
                pdf.cell(25, 8, 'Excess', 1, 0, 'C')
                pdf.cell(15, 8, 'Missing', 1, 0, 'C')
                pdf.ln()
                
                # Second row for "After Loss", "Energy", "Energy", "Info"
                pdf.cell(20, 8, '', 1, 0, 'C')  # Border for Date column
                pdf.cell(25, 8, '', 1, 0, 'C')  # Border for Time column
                pdf.cell(15, 8, '', 1, 0, 'C')  # Border for TOD column
                pdf.cell(25, 8, 'After Loss', 1, 0, 'C')
                pdf.cell(25, 8, 'Energy', 1, 0, 'C')
                pdf.cell(25, 8, 'Energy', 1, 0, 'C')
                pdf.cell(15, 8, 'Info', 1, 0, 'C')
                pdf.ln()
                
                # Third row for units "(kWh)"
                pdf.cell(20, 8, '', 1, 0, 'C')  # Border for Date column
                pdf.cell(25, 8, '', 1, 0, 'C')  # Border for Time column
                pdf.cell(15, 8, '', 1, 0, 'C')  # Border for TOD column
                pdf.cell(25, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(25, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(25, 8, '(kWh)', 1, 0, 'C')
                pdf.cell(15, 8, '', 1, 0, 'C')  # Border for Missing Info column
                pdf.ln()
            
            # Reset font to table data font after headers
            pdf.set_font('Arial', '', 8)  # Single-source table content font size
        
        # Add initial table headers
        add_table_headers()
        
        # Set table content font size based on source type
        is_dual_source = data.get('enable_iex') and data.get('enable_cpp')
        if is_dual_source:
            pdf.set_font('Arial', '', 9)  # Multi-source: table content font size 9
        else:
            pdf.set_font('Arial', '', 8)  # Single-source: table content font size 8
        
        table_complete = False  # Flag to track if table data is finished
        
        # Helper function for proper rounding (≥0.5 rounds up) to avoid scope issues
        def round_excess(value):
            return int(value + 0.5) if value >= 0 else int(value - 0.5)
        
        for idx, row in pdf_data.iterrows():
            # Check if we need a new page (leaving space for summary)
            if pdf.get_y() > 250:  # Near bottom of page
                pdf.add_page()
                # Only add headers if we're still in the table data section
                if not table_complete:
                    add_table_headers()  # Add headers on new page only for table data
            
            if is_dual_source:
                # Sequential adjustment table data with adjusted column widths
                pdf.cell(16, 7, safe_date_str(row['Slot_Date']), 1, 0, 'C')
                pdf.cell(20, 7, format_time(row['Slot_Time']), 1, 0, 'C')  # Increased time column width
                
                tod_cat = row.get('TOD_Category', '')
                pdf.cell(12, 7, tod_cat, 1, 0, 'C')
                
                pdf.cell(18, 7, f"{round_excess(row['Energy_kWh_cons'])}", 1, 0, 'C')
                pdf.cell(18, 7, f"{round_excess(row.get('IEX_After_Loss', 0))}", 1, 0, 'C')
                # Show rounded excess values instead of decimals
                iex_excess_rounded = round_excess(row.get('IEX_Excess', 0))
                pdf.cell(16, 7, f"{iex_excess_rounded}", 1, 0, 'C')
                pdf.cell(18, 7, f"{round_excess(row.get('CPP_After_Loss', 0))}", 1, 0, 'C')
                cpp_excess_rounded = round_excess(row.get('CPP_Excess', 0))
                pdf.cell(16, 7, f"{cpp_excess_rounded}", 1, 0, 'C')
                total_excess_rounded = round_excess(row['Total_Excess'])
                pdf.cell(18, 7, f"{total_excess_rounded}", 1, 0, 'C')
                pdf.cell(12, 7, row.get('Missing_Info', '')[:3], 1, 0, 'C')  # Truncate missing info
            else:
                # Standard table data for single source
                pdf.cell(20, 7, safe_date_str(row['Slot_Date']), 1, 0, 'C')
                pdf.cell(25, 7, format_time(row['Slot_Time']), 1, 0, 'C')  # Increased time column width
                
                tod_cat = row.get('TOD_Category', '')
                pdf.cell(15, 7, tod_cat, 1, 0, 'C')
                
                pdf.cell(25, 7, f"{row['After_Loss']:.2f}", 1, 0, 'C')
                pdf.cell(25, 7, f"{row['Energy_kWh_cons']:.2f}", 1, 0, 'C')
                # Show rounded excess values instead of decimals
                total_excess_rounded = round_excess(row['Total_Excess'])
                pdf.cell(25, 7, f"{total_excess_rounded}", 1, 0, 'C')
                pdf.cell(15, 7, row.get('Missing_Info', '')[:4], 1, 0, 'C')
            
            pdf.ln()
        
        # Mark table as complete - no more headers needed for subsequent pages
        table_complete = True
        
        pdf.ln(2)
        
        # Check if we need a new page for summary (but don't add table headers)
        if pdf.get_y() > 220:  # Need more space for summary
            pdf.add_page()
        
        pdf.ln(2)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'DETAILED CALCULATION SUMMARY:', ln=True)
        pdf.set_font('Arial', '', 11)
        
        # Helper function for proper rounding (≥0.5 rounds up) to match table values
        def round_kwh_summary(value):
            return int(value + 0.5) if value >= 0 else int(value - 0.5)
        
        # Calculate totals
        total_excess = data['total_excess']
        total_consumed = data['total_consumed']
        total_generated_after_loss = data['total_generated_after_loss']
        
        # Enhanced summary for sequential adjustment
        if is_dual_source:
            # Sequential adjustment summary - use rounded totals from table data for precision
            total_iex_before_loss = data.get('merged_all', pdf_data)['IEX_Energy_kWh'].sum()
            total_cpp_before_loss = data.get('merged_all', pdf_data)['CPP_Energy_kWh'].sum()
            total_iex_after_loss = data.get('merged_all', pdf_data)['IEX_After_Loss'].sum()
            total_cpp_after_loss = data.get('merged_all', pdf_data)['CPP_After_Loss'].sum()
            total_iex_excess = data.get('merged_all', pdf_data)['IEX_Excess'].sum()
            total_cpp_excess = data.get('merged_all', pdf_data)['CPP_Excess'].sum()
            
            # Round all values
            total_iex_before_loss_rounded = round_kwh_summary(total_iex_before_loss)
            total_iex_after_loss_rounded = round_kwh_summary(total_iex_after_loss)
            total_cpp_before_loss_rounded = round_kwh_summary(total_cpp_before_loss)
            total_cpp_after_loss_rounded = round_kwh_summary(total_cpp_after_loss)
            total_iex_excess_rounded = round_kwh_summary(total_iex_excess)
            total_cpp_excess_rounded = round_kwh_summary(total_cpp_excess)
            
            pdf.cell(0, 8, f'I.E.X Generation (before T&D loss): {total_iex_before_loss_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'I.E.X Generation (after {data.get("t_and_d_loss", 0)}% T&D loss): {total_iex_after_loss_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'I.E.X Excess Energy (rounded): {total_iex_excess_rounded} kWh', ln=True)
            pdf.ln(3)
            
            pdf.cell(0, 8, f'C.P.P Generation (before T&D loss): {total_cpp_before_loss_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'C.P.P Generation (after {data.get("cpp_t_and_d_loss", 0)}% T&D loss): {total_cpp_after_loss_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'C.P.P Excess Energy (rounded): {total_cpp_excess_rounded} kWh', ln=True)
            pdf.ln(3)
            
            pdf.set_font('Arial', 'B', 11)
            pdf.cell(0, 8, 'TOTAL CALCULATIONS:', ln=True)
            pdf.set_font('Arial', '', 11)
            total_generation_before_rounded = round_kwh_summary(total_iex_before_loss + total_cpp_before_loss)
            total_generation_after_rounded = round_kwh_summary(total_iex_after_loss + total_cpp_after_loss)
            total_consumed_rounded = round_kwh_summary(total_consumed)
            total_excess_rounded = total_iex_excess_rounded + total_cpp_excess_rounded
            
            pdf.cell(0, 8, f'Total Generation (before loss): {total_generation_before_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'Total Generation (after loss): {total_generation_after_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'Total Consumed Energy: {total_consumed_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'Total Excess Energy (rounded): {total_excess_rounded} kWh', ln=True)
        else:
            # Single source summary
            total_excess_rounded = round_kwh_summary(total_excess)
            total_consumed_rounded = round_kwh_summary(total_consumed)
            total_generated_after_loss_rounded = round_kwh_summary(total_generated_after_loss)
            
            if data.get('enable_iex'):
                pdf.cell(0, 8, f'I.E.X Generation (after {data.get("t_and_d_loss", 0)}% T&D loss): {total_generated_after_loss_rounded} kWh', ln=True)
            elif data.get('enable_cpp'):
                pdf.cell(0, 8, f'C.P.P Generation (after {data.get("cpp_t_and_d_loss", 0)}% T&D loss): {total_generated_after_loss_rounded} kWh', ln=True)
            
            pdf.cell(0, 8, f'Total Consumed Energy (after multiplication): {total_consumed_rounded} kWh', ln=True)
            pdf.cell(0, 8, f'Total Excess Energy (rounded): {total_excess_rounded} kWh', ln=True)
        
        pdf.cell(0, 8, f'Unique Days Used (Generated): {data["unique_days_gen"]}', ln=True)
        pdf.cell(0, 8, f'Unique Days Used (Consumed): {data["unique_days_cons"]}', ln=True)
        pdf.cell(0, 8, f'Status: {data["excess_status"]}', ln=True)
        
        # Check if we need a new page for TOD breakdown
        if pdf.get_y() > 220:
            pdf.add_page()
        
        # Add TOD-wise excess energy breakdown
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'TOD-wise Excess Energy Breakdown:', ln=True)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(20, 10, 'TOD', 1)
        pdf.cell(50, 10, 'Excess Energy (kWh)', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 10)
        
        # Calculate TOD-wise excess from the dataframe
        tod_excess = pdf_data.groupby('TOD_Category')['Total_Excess'].sum().reset_index()
        
        # Calculate C category total (sum of C1, C2, C4, C5)
        c_categories = ['C1', 'C2', 'C4', 'C5']
        c_total = 0
        tod_values = {}
        
        for _, row in tod_excess.iterrows():
            category = row['TOD_Category']
            excess_rounded = round_kwh_summary(row['Total_Excess'])
            tod_values[category] = excess_rounded
            if category in c_categories:
                c_total += excess_rounded
        
        # Display C total first
        pdf.cell(20, 10, 'C', 1)
        pdf.cell(50, 10, f"{c_total}", 1)
        pdf.ln()
        
        # Display individual categories
        for category in ['C1', 'C2', 'C4', 'C5']:
            if category in tod_values:
                pdf.cell(20, 10, category, 1)
                pdf.cell(50, 10, f"{tod_values[category]}", 1)
                pdf.ln()
        
        if 'Unknown' in tod_values:
            pdf.cell(20, 10, 'Unknown', 1)
            pdf.cell(50, 10, f"{tod_values['Unknown']}", 1)
            pdf.ln()
        
        # Add financial calculations (using pre-calculated values from data processing)
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Financial Calculations:', ln=True)
        pdf.set_font('Arial', '', 10)
        
        # Use pre-calculated financial values if available, otherwise calculate on the fly
        additional_surcharge = 0.0
        additional_surcharge_breakdown = []
        if 'total_excess_financial_rounded' in data:
            total_excess_financial_rounded = data['total_excess_financial_rounded']
            base_rate = data['base_rate']
            c1_c2_rate = data.get('tariff_c1_c2_rate', c1_c2_rate)
            c5_rate = data.get('tariff_c5_rate', c5_rate)
            cross_subsidy_rate = data.get('tariff_cross_subsidy_rate', cross_subsidy_rate)
            wheeling_rate = data.get('tariff_wheeling_rate', wheeling_rate)
            base_amount = data['base_amount']
            c1_c2_excess = data['c1_c2_excess']
            c1_c2_additional = data['c1_c2_additional']
            c5_excess = data['c5_excess']
            c5_additional = data['c5_additional']
            total_amount = data['total_amount']
            etax = data['etax']
            total_with_etax = data['total_with_etax']
            iex_excess_financial = data['iex_excess_financial']
            etax_on_iex = data['etax_on_iex']
            cross_subsidy_surcharge = data['cross_subsidy_surcharge']
            additional_surcharge = data.get('additional_surcharge', 0.0)
            additional_surcharge_breakdown = data.get('additional_surcharge_breakdown', [])
            wheeling_charges = data['wheeling_charges']
            wheeling_reference_kwh = data.get('wheeling_reference_kwh', 0.0)
            wheeling_combined_kwh = data.get('wheeling_combined_kwh', total_excess_financial_rounded + wheeling_reference_kwh)
            wheeling_reduction_kwh = data.get('wheeling_reduction_kwh', wheeling_combined_kwh * WHEELING_PERCENT)
            final_amount = data['final_amount']
            final_amount_rounded = data['final_amount_rounded']
        else:
            # Calculate on the fly (fallback)
            total_excess_financial_rounded = round_kwh_summary(total_excess)
            base_rate = tariff_rates['base_rate']
            c1_c2_rate = tariff_rates['c1_c2_rate']
            c5_rate = tariff_rates['c5_rate']
            cross_subsidy_rate = tariff_rates['cross_subsidy_rate']
            wheeling_rate = tariff_rates['wheeling_rate']
            base_amount = total_excess_financial_rounded * base_rate
            c1_c2_excess = tod_values.get('C1', 0) + tod_values.get('C2', 0)
            c1_c2_additional = c1_c2_excess * c1_c2_rate
            c5_excess = tod_values.get('C5', 0)
            c5_additional = c5_excess * c5_rate
            total_amount = base_amount + c1_c2_additional + c5_additional
            etax = total_amount * 0.05
            total_with_etax = total_amount + etax

            # Calculate IEX excess for specific charges
            iex_excess_financial_raw = 0
            if 'IEX_Excess' in pdf_data.columns:
                iex_excess_financial_raw = pdf_data['IEX_Excess'].sum()
            elif data.get('enable_iex') and not data.get('enable_cpp'):
                iex_excess_financial_raw = total_excess

            iex_excess_financial = round_kwh_summary(iex_excess_financial_raw)
            etax_on_iex = total_excess_financial_rounded * 0.1
            cross_subsidy_surcharge = iex_excess_financial * cross_subsidy_rate

            additional_surcharge, additional_surcharge_breakdown, additional_surcharge_rate, additional_surcharge_period_label, additional_surcharge_note = calculate_monthly_additional_surcharge(
                data.get('month'),
                data.get('year'),
                iex_excess_financial_raw,
                iex_excess_financial,
            )

            wheeling_components = compute_wheeling_components(
                total_excess_financial_rounded,
                data.get('t_and_d_loss', 0.0),
            )
            wheeling_reference_kwh = wheeling_components['reference_kwh']
            wheeling_combined_kwh = wheeling_components['combined_kwh']
            wheeling_reduction_kwh = wheeling_components['reduction_kwh']
            wheeling_adjusted_kwh = wheeling_components['adjusted_kwh']
            wheeling_charges = wheeling_adjusted_kwh * wheeling_rate

            final_amount = total_with_etax - (etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge)
            final_amount_rounded = math.ceil(final_amount)

        # Display the financial calculations with proper formatting
        pdf.cell(0, 8, f"1. Base Rate: Total Excess Energy ({total_excess_financial_rounded} kWh) x Rs.{base_rate:.4f} = Rs.{base_amount:.2f}", ln=True)
        pdf.cell(0, 8, f"2. C1+C2 Additional: Excess in C1+C2 ({c1_c2_excess} kWh) x Rs.{c1_c2_rate:.4f} = Rs.{c1_c2_additional:.2f}", ln=True)
        pdf.cell(0, 8, f"3. C5 Additional: Excess in C5 ({c5_excess} kWh) x Rs.{c5_rate:.4f} = Rs.{c5_additional:.2f}", ln=True)
        pdf.cell(0, 8, f"4. Partial Total: Rs.{base_amount:.2f} + Rs.{c1_c2_additional:.2f} + Rs.{c5_additional:.2f} = Rs.{total_amount:.2f}", ln=True)
        pdf.cell(0, 8, f"5. E-Tax (5% of Partial Total): Rs.{total_amount:.2f} x 0.05 = Rs.{etax:.2f}", ln=True)
        pdf.cell(0, 8, f"6. Subtotal with E-Tax: Rs.{total_amount:.2f} + Rs.{etax:.2f} = Rs.{total_with_etax:.2f}", ln=True)
        pdf.cell(0, 8, f"7. Less: E-Tax on IEX: Total Excess ({total_excess_financial_rounded} kWh) x Rs.0.1 = Rs.{etax_on_iex:.2f}", ln=True)
        pdf.cell(0, 8, f"8. Less: Cross Subsidy Surcharge: IEX Excess ({iex_excess_financial} kWh) x Rs.{cross_subsidy_rate:.4f} = Rs.{cross_subsidy_surcharge:.2f}", ln=True)
        pdf.cell(0, 8, f"8a. Less: Additional Surcharge (IEX): Rs.{additional_surcharge:.2f}", ln=True)
        pdf.cell(0, 8, f"9. Wheeling Reference: Total Excess ({total_excess_financial_rounded} kWh) + Rounded Loss ({wheeling_reference_kwh:.2f} kWh) = {wheeling_combined_kwh:.2f} kWh", ln=True)
        pdf.cell(0, 8, f"9a. Less 2.34%: {wheeling_combined_kwh:.2f} kWh × 2.34% = {wheeling_reduction_kwh:.2f} kWh", ln=True)
        pdf.cell(0, 8, f"9b. Wheeling Charges: ({wheeling_combined_kwh:.2f} - {wheeling_reduction_kwh:.2f}) kWh × Rs.{wheeling_rate:.4f} = Rs.{wheeling_charges:.2f}", ln=True)
        # If breakdown available, print details per date-range
        if additional_surcharge_breakdown:
            for entry in additional_surcharge_breakdown:
                if isinstance(entry, dict):
                    period_text = entry.get('period_label') or entry.get('window_label') or 'Selected Period'
                    rate_value = entry.get('rate', 0.0)
                    kwh_value = entry.get('kwh', entry.get('raw_kwh', 0))
                    amount_value = entry.get('amount', 0.0)
                    note_value = entry.get('note', '')
                    pdf.cell(0, 6, f"    - {period_text}: {kwh_value} kWh x Rs.{rate_value:.2f} per kWh = Rs.{amount_value:.4f} ({note_value})", ln=True)
                else:
                    try:
                        s_start, s_end, s_excess, s_paise, s_comp, s_note = entry
                        pdf.cell(0, 6, f"    - {s_start} to {s_end}: {s_excess} kWh x Rs.{s_paise} per kWh = Rs.{s_comp:.4f} ({s_note})", ln=True)
                    except Exception:
                        continue
            
        # Calculate deductions total for clarity (include Additional Surcharge)
        deductions_total = etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge
        pdf.cell(0, 8, f"10a. Total Amount to be Collected - Step 1:", ln=True)
        pdf.cell(0, 8, f"     Rs.{total_with_etax:.2f} - (Rs.{etax_on_iex:.2f} + Rs.{cross_subsidy_surcharge:.2f} + Rs.{wheeling_charges:.2f} + Rs.{additional_surcharge:.2f})", ln=True)
        pdf.cell(0, 8, f"10b. Total Amount to be Collected - Step 2:", ln=True)
        pdf.cell(0, 8, f"     Rs.{total_with_etax:.2f} - Rs.{deductions_total:.2f} = Rs.{final_amount:.2f}", ln=True)

        # Round up final amount to next highest value
        pdf.set_font('Arial', 'B', 10)  # Consistent with table data font size
        pdf.cell(0, 8, f"11. Final Amount (Rounded Up): Rs.{final_amount_rounded}", ln=True)

        # Generate PDF bytes
        pdf_output = io.BytesIO()
        pdf_bytes = pdf.output(dest='S')
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin1')
        pdf_output.write(pdf_bytes)
        pdf_output.seek(0)
        return pdf_output.getvalue()

    except Exception as e:
        st.error(f"Error generating detailed PDF: {str(e)}")
        return None

def generate_daywise_pdf(data, pdf_data):
    """Generate day-wise summary PDF"""
    try:
        # Defensive defaults for additional surcharge to satisfy static analysis
        additional_surcharge = data.get('additional_surcharge', 0.0)
        additional_surcharge_breakdown = data.get('additional_surcharge_breakdown', [])

        # Resolve tariff metadata for rate lookups
        tariff_selection = data.get('tariff_selection', TARIFF_OPTIONS[0])
        tariff_rates = data.get('tariff_rates') or resolve_tariff_rates(
            tariff_selection,
            data.get('month'),
            data.get('year'),
        )
        base_rate = data.get('base_rate', tariff_rates['base_rate'])
        c1_c2_rate = data.get('tariff_c1_c2_rate', tariff_rates['c1_c2_rate'])
        c5_rate = data.get('tariff_c5_rate', tariff_rates['c5_rate'])
        cross_subsidy_rate = data.get('tariff_cross_subsidy_rate', tariff_rates['cross_subsidy_rate'])
        wheeling_rate = data.get('tariff_wheeling_rate', tariff_rates['wheeling_rate'])

        def round_kwh_summary(value):
            try:
                value = float(value)
            except Exception:
                value = 0.0
            return int(value + 0.5) if value >= 0 else int(value - 0.5)
        
        pdf = AuthorPDF()
        pdf.set_margins(20, 20, 20)
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()
        
        # FIRST PAGE - DESCRIPTION AND INFORMATION ONLY
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 15, 'Energy Adjustment Day-wise Summary Report', ln=True, align='C')
        pdf.ln(10)
        
        # Consumer Information Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Consumer Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f"Consumer Number: {data['consumer_number']}", ln=True)
        pdf.cell(0, 8, f"Consumer Name: {data['consumer_name']}", ln=True)
        pdf.cell(0, 8, f"Multiplication Factor (Consumed Energy): {data.get('multiplication_factor', 1)}", ln=True)
        pdf.cell(0, 8, f"Tariff: {tariff_selection} ({tariff_rates.get('window_label', 'Latest')})", ln=True)
        pdf.ln(5)
        
        # Technical Parameters Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Technical Parameters:', ln=True)
        pdf.set_font('Arial', '', 12)
        
        if data.get('enable_iex') and data.get('enable_cpp'):
            pdf.cell(0, 8, f"I.E.X T&D Loss (%): {data.get('t_and_d_loss', 0)}", ln=True)
            pdf.cell(0, 8, f"C.P.P T&D Loss (%): {data.get('cpp_t_and_d_loss', 0)}", ln=True)
        elif data.get('enable_iex'):
            pdf.cell(0, 8, f"I.E.X T&D Loss (%): {data.get('t_and_d_loss', 0)}", ln=True)
        elif data.get('enable_cpp'):
            pdf.cell(0, 8, f"C.P.P T&D Loss (%): {data.get('cpp_t_and_d_loss', 0)}", ln=True)
        pdf.ln(5)
        
        # Report Information Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Report Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f"Report Generated: {datetime.now().strftime('%d/%m/%Y at %H:%M:%S')}", ln=True)
        if data.get('month') and data.get('year'):
            pdf.cell(0, 8, f"Report Period: {data['month']}/{data['year']}", ln=True)
        pdf.ln(10)
        
        # Report Description Section
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Report Description:', ln=True)
        pdf.set_font('Arial', '', 11)
        
        description_text = [
            "This day-wise summary report consolidates energy adjustment calculations",
            "into daily totals, showing the overall energy balance for each day",
            "of the selected month. The report includes all days in the month,",
            "even those with no energy transactions, for complete visibility."
        ]
        
        for line in description_text:
            pdf.cell(0, 6, line, ln=True)
        
        pdf.ln(10)
        
        # Add page break before tables
        pdf.add_page()
        
        # SECOND PAGE AND BEYOND - TABLES START HERE
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 10, 'Day-wise Summary Table', ln=True, align='C')
        pdf.ln(10)
        
        # Table headers sized to respect left/right margins
        usable_width = pdf.w - pdf.l_margin - pdf.r_margin
        date_col_width = usable_width * 0.22
        other_col_width = (usable_width - date_col_width) / 3
        table_columns = [
            ('Date', date_col_width),
            ('Total Gen. After Loss', other_col_width),
            ('Total Consumed', other_col_width),
            ('Total Excess', other_col_width),
        ]

        pdf.set_font('Arial', 'B', 9)
        for header, width in table_columns:
            pdf.cell(width, 10, header, 1)
        pdf.ln()
        
        # Calculate day-wise data
        if data.get('enable_iex') and data.get('enable_cpp'):
            daywise = pdf_data.groupby('Slot_Date').agg({
                'IEX_After_Loss': 'sum',
                'CPP_After_Loss': 'sum',
                'Energy_kWh_cons': 'sum',
                'Total_Excess': 'sum'
            }).reset_index()
            daywise['Total_After_Loss'] = daywise['IEX_After_Loss'] + daywise['CPP_After_Loss']
        else:
            daywise = pdf_data.groupby('Slot_Date').agg({
                'After_Loss': 'sum',
                'Energy_kWh_cons': 'sum',
                'Total_Excess': 'sum'
            }).reset_index()
            daywise['Total_After_Loss'] = daywise['After_Loss']
        
        pdf.set_font('Arial', '', 8)
        for _, row in daywise.iterrows():
            pdf.cell(date_col_width, 10, str(row['Slot_Date']), 1)
            pdf.cell(other_col_width, 10, f"{row['Total_After_Loss']:.4f}", 1)
            pdf.cell(other_col_width, 10, f"{row['Energy_kWh_cons']:.4f}", 1)

            # Round excess values for display
            total_excess_rounded = int(row['Total_Excess'] + 0.5) if row['Total_Excess'] >= 0 else int(row['Total_Excess'] - 0.5)
            pdf.cell(other_col_width, 10, f"{total_excess_rounded}", 1)
            pdf.ln()
        
        # Add same calculation summary as detailed PDF
        pdf.ln(10)
        pdf.set_font('Arial', 'B', 12)
        pdf.cell(0, 10, 'DETAILED CALCULATION SUMMARY:', ln=True)
        pdf.set_font('Arial', '', 11)
        
        # Helper function for proper rounding
        def round_kwh(value):
            return int(value + 0.5) if value >= 0 else int(value - 0.5)
        
        # Include same financial calculations as detailed PDF
        total_excess = data['total_excess']
        total_excess_rounded = round_kwh(total_excess)
        total_consumed = data['total_consumed']
        total_consumed_rounded = round_kwh(total_consumed)
        total_generated_after_loss = data['total_generated_after_loss']
        total_generated_after_loss_rounded = round_kwh(total_generated_after_loss)
        
        # Basic summary
        pdf.cell(0, 8, f'Total Generated (after loss): {total_generated_after_loss_rounded} kWh', ln=True)
        pdf.cell(0, 8, f'Total Consumed Energy (after multiplication): {total_consumed_rounded} kWh', ln=True)
        pdf.cell(0, 8, f'Total Excess Energy (rounded): {total_excess_rounded} kWh', ln=True)
        pdf.cell(0, 8, f'Unique Days Used (Generated): {data["unique_days_gen"]}', ln=True)
        pdf.cell(0, 8, f'Unique Days Used (Consumed): {data["unique_days_cons"]}', ln=True)
        pdf.cell(0, 8, f'Status: {data["excess_status"]}', ln=True)
        
        # Check if we need a new page for TOD breakdown
        if pdf.get_y() > 220:
            pdf.add_page()
        
        # Add TOD-wise excess energy breakdown
        pdf.ln(5)
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'TOD-wise Excess Energy Breakdown:', ln=True)
        pdf.set_font('Arial', 'B', 10)
        pdf.cell(20, 10, 'TOD', 1)
        pdf.cell(50, 10, 'Excess Energy (kWh)', 1)
        pdf.ln()
        
        pdf.set_font('Arial', '', 10)
        
        # Calculate TOD-wise excess from the dataframe
        tod_excess = pdf_data.groupby('TOD_Category')['Total_Excess'].sum().reset_index()
        
        # Calculate C category total (sum of C1, C2, C4, C5)
        c_categories = ['C1', 'C2', 'C4', 'C5']
        c_total = 0
        tod_values = {}
        
        for _, row in tod_excess.iterrows():
            category = row['TOD_Category']
            excess_rounded = round_kwh(row['Total_Excess'])
            tod_values[category] = excess_rounded
            if category in c_categories:
                c_total += excess_rounded
        
        # Display C total first
        pdf.cell(20, 10, 'C', 1)
        pdf.cell(50, 10, f"{c_total}", 1)
        pdf.ln()
        
        # Display individual categories
        for category in ['C1', 'C2', 'C4', 'C5']:
            if category in tod_values:
                pdf.cell(20, 10, category, 1)
                pdf.cell(50, 10, f"{tod_values[category]}", 1)
                pdf.ln()
        
        if 'Unknown' in tod_values:
            pdf.cell(20, 10, 'Unknown', 1)
            pdf.cell(50, 10, f"{tod_values['Unknown']}", 1)
            pdf.ln()
        
        # Add financial calculations (using pre-calculated values from data processing)
        pdf.add_page()
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Financial Calculations:', ln=True)
        pdf.set_font('Arial', '', 10)

        if 'total_excess_financial_rounded' in data:
            total_excess_financial_rounded = data['total_excess_financial_rounded']
            base_rate = data['base_rate']
            c1_c2_rate = data.get('tariff_c1_c2_rate', c1_c2_rate)
            c5_rate = data.get('tariff_c5_rate', c5_rate)
            cross_subsidy_rate = data.get('tariff_cross_subsidy_rate', cross_subsidy_rate)
            wheeling_rate = data.get('tariff_wheeling_rate', wheeling_rate)
            base_amount = data['base_amount']
            c1_c2_excess = data['c1_c2_excess']
            c1_c2_additional = data['c1_c2_additional']
            c5_excess = data['c5_excess']
            c5_additional = data['c5_additional']
            total_amount = data['total_amount']
            etax = data['etax']
            total_with_etax = data['total_with_etax']
            iex_excess_financial = data['iex_excess_financial']
            etax_on_iex = data['etax_on_iex']
            cross_subsidy_surcharge = data['cross_subsidy_surcharge']
            wheeling_charges = data['wheeling_charges']
            wheeling_reference_kwh = data.get('wheeling_reference_kwh', 0.0)
            wheeling_combined_kwh = data.get('wheeling_combined_kwh', total_excess_financial_rounded + wheeling_reference_kwh)
            wheeling_reduction_kwh = data.get('wheeling_reduction_kwh', wheeling_combined_kwh * WHEELING_PERCENT)
            wheeling_adjusted_kwh = data.get('wheeling_adjusted_kwh', wheeling_combined_kwh - wheeling_reduction_kwh)
            final_amount = data['final_amount']
            final_amount_rounded = data['final_amount_rounded']
        else:
            total_excess_financial_rounded = round_kwh_summary(total_excess)
            base_rate = tariff_rates['base_rate']
            c1_c2_rate = tariff_rates['c1_c2_rate']
            c5_rate = tariff_rates['c5_rate']
            cross_subsidy_rate = tariff_rates['cross_subsidy_rate']
            wheeling_rate = tariff_rates['wheeling_rate']
            base_amount = total_excess_financial_rounded * base_rate
            c1_c2_excess = tod_values.get('C1', 0) + tod_values.get('C2', 0)
            c1_c2_additional = c1_c2_excess * c1_c2_rate
            c5_excess = tod_values.get('C5', 0)
            c5_additional = c5_excess * c5_rate
            total_amount = base_amount + c1_c2_additional + c5_additional
            etax = total_amount * 0.05
            total_with_etax = total_amount + etax

            iex_excess_financial_raw = 0
            if data.get('enable_iex') and 'IEX_Excess' in pdf_data.columns:
                iex_excess_financial_raw = pdf_data['IEX_Excess'].sum()
            elif data.get('enable_iex') and not data.get('enable_cpp'):
                iex_excess_financial_raw = total_excess

            iex_excess_financial = round_kwh_summary(iex_excess_financial_raw)
            etax_on_iex = total_excess_financial_rounded * 0.1
            cross_subsidy_surcharge = iex_excess_financial * cross_subsidy_rate

            additional_surcharge, additional_surcharge_breakdown, additional_surcharge_rate, additional_surcharge_period_label, additional_surcharge_note = calculate_monthly_additional_surcharge(
                data.get('month'),
                data.get('year'),
                iex_excess_financial_raw,
                iex_excess_financial,
            )

            wheeling_components = compute_wheeling_components(
                total_excess_financial_rounded,
                data.get('t_and_d_loss', 0.0),
            )
            wheeling_reference_kwh = wheeling_components['reference_kwh']
            wheeling_combined_kwh = wheeling_components['combined_kwh']
            wheeling_reduction_kwh = wheeling_components['reduction_kwh']
            wheeling_adjusted_kwh = wheeling_components['adjusted_kwh']
            wheeling_charges = wheeling_adjusted_kwh * wheeling_rate

            final_amount = total_with_etax - (etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge)
            final_amount_rounded = math.ceil(final_amount)

        line_gap = 1.8

        def add_spaced_line(text, bold=False):
            pdf.set_font('Arial', 'B' if bold else '', 10)
            pdf.cell(0, 8, text, ln=True)
            pdf.ln(line_gap)

        add_spaced_line(f"1. Base Rate: Total Excess Energy ({total_excess_financial_rounded} kWh) x Rs.{base_rate:.4f} = Rs.{base_amount:.2f}")
        add_spaced_line(f"2. C1+C2 Additional: Excess in C1+C2 ({c1_c2_excess} kWh) x Rs.{c1_c2_rate:.4f} = Rs.{c1_c2_additional:.2f}")
        add_spaced_line(f"3. C5 Additional: Excess in C5 ({c5_excess} kWh) x Rs.{c5_rate:.4f} = Rs.{c5_additional:.2f}")
        add_spaced_line(f"4. Partial Total: Rs.{total_amount:.2f}")
        add_spaced_line(f"5. E-Tax (5%): Rs.{etax:.2f}")
        add_spaced_line(f"6. Subtotal with E-Tax: Rs.{total_with_etax:.2f}")
        add_spaced_line(f"7. Less: E-Tax on IEX: Rs.{etax_on_iex:.2f}")
        add_spaced_line(f"8. Less: Cross Subsidy Surcharge: {iex_excess_financial} kWh x Rs.{cross_subsidy_rate:.4f} = Rs.{cross_subsidy_surcharge:.2f}")
        add_spaced_line(f"8a. Less: Additional Surcharge (IEX): Rs.{additional_surcharge:.2f}")
        if additional_surcharge_breakdown:
            for entry in additional_surcharge_breakdown:
                if isinstance(entry, dict):
                    period_text = entry.get('period_label') or entry.get('window_label') or 'Selected Period'
                    rate_value = entry.get('rate', 0.0)
                    kwh_value = entry.get('kwh', entry.get('raw_kwh', 0))
                    amount_value = entry.get('amount', 0.0)
                    note_value = entry.get('note', '')
                    pdf.cell(0, 6, f"    - {period_text}: {kwh_value} kWh x Rs.{rate_value:.2f} per kWh = Rs.{amount_value:.4f} ({note_value})", ln=True)
                    pdf.ln(0.8)
                else:
                    try:
                        s_start, s_end, s_excess, s_paise, s_comp, s_note = entry
                        pdf.cell(0, 6, f"    - {s_start} to {s_end}: {s_excess} kWh x Rs.{s_paise} per kWh = Rs.{s_comp:.4f} ({s_note})", ln=True)
                        pdf.ln(0.8)
                    except Exception:
                        continue

        add_spaced_line(
            f"9. Wheeling Reference: Total Excess ({total_excess_financial_rounded} kWh) + Rounded Loss ({wheeling_reference_kwh:.2f} kWh) = {wheeling_combined_kwh:.2f} kWh"
        )
        add_spaced_line(
            f"9a. Less 2.34%: {wheeling_combined_kwh:.2f} kWh × 2.34% = {wheeling_reduction_kwh:.2f} kWh"
        )
        add_spaced_line(
            f"9b. Wheeling Charges: ({wheeling_combined_kwh:.2f} - {wheeling_reduction_kwh:.2f}) kWh × Rs.{wheeling_rate:.4f} = Rs.{wheeling_charges:.2f}"
        )

        deductions_total = etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge

        add_spaced_line(f"10. Final Amount: Rs.{total_with_etax:.2f} - Rs.{deductions_total:.2f} = Rs.{final_amount:.2f}", bold=True)
        add_spaced_line(f"11. Final Amount (Rounded Up): Rs.{final_amount_rounded}", bold=True)

        # Generate PDF bytes
        pdf_output = io.BytesIO()
        pdf_bytes = pdf.output(dest='S')
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin1')
        pdf_output.write(pdf_bytes)
        pdf_output.seek(0)
        return pdf_output.getvalue()

    except Exception as e:
        st.error(f"Error generating daywise PDF: {str(e)}")
        return None

def generate_simple_pdf(data, pdf_type="excess"):
    """Generate a simple PDF report"""
    try:
        pdf = AuthorPDF()
        pdf.set_margins(20, 20, 20)
        pdf.set_auto_page_break(auto=True, margin=20)
        pdf.add_page()
        
        # Title
        pdf.set_font('Arial', 'B', 16)
        pdf.cell(0, 15, 'Energy Adjustment Report', ln=True, align='C')
        pdf.ln(10)
        
        # Consumer Information
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Consumer Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f'Consumer Number: {data["consumer_number"]}', ln=True)
        pdf.cell(0, 8, f'Consumer Name: {data["consumer_name"]}', ln=True)
        pdf.cell(0, 8, f'Multiplication Factor: {data["multiplication_factor"]}', ln=True)
        tariff_selection = data.get('tariff_selection', TARIFF_OPTIONS[0])
        tariff_rates = data.get('tariff_rates') or resolve_tariff_rates(
            tariff_selection,
            data.get('month'),
            data.get('year'),
        )
        pdf.cell(0, 8, f'Tariff: {tariff_selection} ({tariff_rates.get("window_label", "Latest")})', ln=True)
        pdf.ln(5)
        
        # Technical Parameters
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Technical Parameters:', ln=True)
        pdf.set_font('Arial', '', 12)
        
        if data['enable_iex'] and data['enable_cpp']:
            pdf.cell(0, 8, f'I.E.X T&D Loss (%): {data["t_and_d_loss"]}', ln=True)
            pdf.cell(0, 8, f'C.P.P T&D Loss (%): {data["cpp_t_and_d_loss"]}', ln=True)
        elif data['enable_iex']:
            pdf.cell(0, 8, f'I.E.X T&D Loss (%): {data["t_and_d_loss"]}', ln=True)
        elif data['enable_cpp']:
            pdf.cell(0, 8, f'C.P.P T&D Loss (%): {data["cpp_t_and_d_loss"]}', ln=True)
        pdf.ln(5)
        
        # Summary
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Summary:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f'Total Generated (After Loss): {data["total_generated_after_loss"]:.2f} kWh', ln=True)
        pdf.cell(0, 8, f'Total Consumed: {data["total_consumed"]:.2f} kWh', ln=True)
        pdf.cell(0, 8, f'Total Excess: {data["total_excess"]:.2f} kWh', ln=True)
        pdf.cell(0, 8, f'Status: {data["excess_status"]}', ln=True)
        pdf.ln(5)
        
        # Report Info
        pdf.set_font('Arial', 'B', 14)
        pdf.cell(0, 10, 'Report Information:', ln=True)
        pdf.set_font('Arial', '', 12)
        pdf.cell(0, 8, f'Report Type: {pdf_type.title()} Slots', ln=True)
        pdf.cell(0, 8, f'Generated: {datetime.now().strftime("%d/%m/%Y at %H:%M:%S")}', ln=True)
        if data.get('auto_detect_info'):
            pdf.cell(0, 8, f'Period: {data["auto_detect_info"]}', ln=True)
        
        # Generate PDF bytes
        pdf_bytes = pdf.output(dest='S')
        if isinstance(pdf_bytes, str):
            pdf_bytes = pdf_bytes.encode('latin1')
        
        return pdf_bytes
        
    except Exception as e:
        st.error(f"Error generating PDF: {str(e)}")
        return None

# First, get the checkboxes outside the form for immediate response
st.header("Input Parameters")

# PDF Output Options
st.subheader("PDF Output Options")
col1, col2, col3 = st.columns(3)
with col1:
    show_excess_only = st.checkbox("Show only slots with excess (loss)", value=True, key="show_excess_only")
with col2:
    show_all_slots = st.checkbox("Show all slots (15-min slot-wise)", key="show_all_slots")
with col3:
    show_daywise = st.checkbox("Show day-wise summary (all days in month)", key="show_daywise")

# Generation Sources Configuration (outside form for immediate response)
st.subheader("Generation Sources Configuration")
col1, col2 = st.columns(2)
with col1:
    enable_iex = st.checkbox("Enable I.E.X Generation", help="Check this to enable I.E.X generation data upload", key="enable_iex")
with col2:
    enable_cpp = st.checkbox("Enable C.P.P Generation", help="Check this to enable C.P.P generation data upload", key="enable_cpp")

# Main form for file uploads and other inputs
with st.form("energy_adjustment_form"):
    # Consumer Information
    st.subheader("Consumer Information")
    col1, col2 = st.columns(2)
    with col1:
        consumer_number = st.text_input("Consumer Number", help="Enter the consumer number")
    with col2:
        consumer_name = st.text_input("Consumer Name", help="Enter the consumer name")
    
    # I.E.X Generation (conditional based on checkbox outside form)
    generated_files = []
    t_and_d_loss = 0.0
    if enable_iex:
        st.subheader("I.E.X Generation Settings")
        generated_files = st.file_uploader(
            "Generated Energy Excel Files I.E.X (MW) From SLDC",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Select one or more Excel files containing I.E.X generation data",
            key="iex_files"
        )
        t_and_d_loss = st.number_input(
            "T&D Loss (%) for I.E.X",
            min_value=0.0,
            max_value=100.0,
            step=0.01,
            format="%.2f",
            help="Enter the transmission and distribution loss percentage for I.E.X",
            key="iex_td_loss",
            value=0.0
        )
    
    # C.P.P Generation (conditional based on checkbox outside form)
    cpp_files = []
    cpp_t_and_d_loss = 0.0
    if enable_cpp:
        st.subheader("C.P.P Generation Settings")
        cpp_files = st.file_uploader(
            "Generated Energy Excel Files C.P.P (MW) From SLDC",
            type=['xlsx', 'xls'],
            accept_multiple_files=True,
            help="Select one or more Excel files containing C.P.P generation data",
            key="cpp_files"
        )
        cpp_t_and_d_loss = st.number_input(
            "T&D Loss (%) for C.P.P",
            min_value=0.0,
            max_value=100.0,
            step=0.01,
            format="%.2f",
            help="Enter the transmission and distribution loss percentage for C.P.P",
            key="cpp_td_loss",
            value=0.0
        )
    
    # Consumed Energy
    st.subheader("Consumed Energy")
    consumed_files = st.file_uploader(
        "Consumed Energy Excel Files (kWh) From MRT",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="Select one or more Excel files containing consumed energy data"
    )
    
    multiplication_factor = st.number_input(
        "Multiplication Factor (for Consumed Energy)",
        min_value=0.01,
        step=0.01,
        format="%.2f",
        help="Factor to multiply consumed energy values"
    )

    tariff_selection = st.selectbox(
        "Tariff Selection",
        options=TARIFF_OPTIONS,
        index=0,
        help="Choose the applicable tariff so that all dependent rates update automatically"
    )
    
    # Date Filters
    st.subheader("Date Filters")
    col1, col2, col3 = st.columns(3)
    with col1:
        date_filter = st.text_input(
            "Date (optional, dd/mm/yyyy)",
            placeholder="e.g. 10/10/2024",
            help="Filter data for a specific date"
        )
    with col2:
        month = st.selectbox(
            "Month (optional)",
            options=[""] + [str(i) for i in range(1, 13)],
            format_func=lambda x: "" if x == "" else datetime(2000, int(x), 1).strftime("%B")
        )
    with col3:
        year = st.number_input(
            "Year (optional)",
            min_value=2000,
            max_value=2100,
            value=None,
            help="Enter the year for filtering data"
        )
    
    auto_detect_month = st.checkbox(
        "Auto-detect month and year from files",
        value=True,
        help="Automatically detect the month and year from uploaded files"
    )
    
    # Debug information inside the form
    st.write("---")
    st.subheader("🔍 Current Configuration")
    col1, col2, col3 = st.columns(3)
    
    with col1:
        st.write("**I.E.X Status:**")
        st.write(f"- Enabled: {enable_iex}")
        if enable_iex:
            st.write(f"- Files: {len(generated_files) if generated_files else 0}")
            st.write(f"- T&D Loss: {t_and_d_loss}%")
    
    with col2:
        st.write("**C.P.P Status:**")
        st.write(f"- Enabled: {enable_cpp}")
        if enable_cpp:
            st.write(f"- Files: {len(cpp_files) if cpp_files else 0}")
            st.write(f"- T&D Loss: {cpp_t_and_d_loss}%")
    
    with col3:
        st.write("**Consumption & Billing:**")
        st.write(f"- Files: {len(consumed_files) if consumed_files else 0}")
        st.write(f"- Factor: {multiplication_factor}")
        st.write(f"- Tariff: {tariff_selection}")
    
    # Submit button
    submitted = st.form_submit_button("Generate PDF Report", type="primary")

# Remove debug information from outside the form

# Validation and processing
if submitted:
    # Clear previous errors
    st.session_state.error_message = None
    
    # Validation
    errors = []
    
    # Check PDF options
    if not (show_excess_only or show_all_slots or show_daywise):
        errors.append("Please select at least one PDF output option.")
    
    # Check consumer information
    if not consumer_number:
        errors.append("Consumer Number is required.")
    if not consumer_name:
        errors.append("Consumer Name is required.")
    
    # Check generation sources
    if not enable_iex and not enable_cpp:
        errors.append("Please enable at least one generation source (I.E.X or C.P.P).")
    
    # Check I.E.X validation
    if enable_iex:
        if not generated_files or len(generated_files) == 0:
            errors.append("Please select I.E.X generation files since I.E.X is enabled.")
        # Remove T&D loss validation to allow 0% loss
    
    # Check C.P.P validation
    if enable_cpp:
        if not cpp_files or len(cpp_files) == 0:
            errors.append("Please select C.P.P generation files since C.P.P is enabled.")
        # Remove T&D loss validation to allow 0% loss
    
    # Check consumed files
    if not consumed_files:
        errors.append("No consumed energy Excel files were uploaded.")
    
    # Check multiplication factor
    if multiplication_factor is None or multiplication_factor <= 0:
        errors.append("Multiplication factor must be greater than 0.")
    
    # Display errors if any
    if errors:
        st.session_state.error_message = "\n".join(errors)
    else:
        # Process the data
        st.success("Validation passed! Processing data...")
        
        try:
            # Process data with progress tracking
            with st.spinner("Processing uploaded files..."):
                result = process_energy_data(
                    generated_files, cpp_files, consumed_files,
                    enable_iex, enable_cpp, t_and_d_loss, cpp_t_and_d_loss,
                    consumer_number, consumer_name, multiplication_factor, tariff_selection,
                    auto_detect_month, month, year, date_filter
                )
                
                if result['success']:
                    st.session_state.processed_data = result['data']
                    st.success("Data processed successfully!")
                else:
                    st.session_state.error_message = result['error']
            
        except Exception as e:
            st.session_state.error_message = f"Error processing data: {str(e)}"

# Display errors
if st.session_state.error_message:
    st.error(st.session_state.error_message)

# Display results if data has been processed
if st.session_state.processed_data:
    st.header("📊 Processing Results")
    
    data = st.session_state.processed_data
    
    # Display summary information
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        st.metric("Total Generated (After Loss)", f"{data['total_generated_after_loss']:.2f} kWh")
    with col2:
        st.metric("Total Consumed", f"{data['total_consumed']:.2f} kWh")
    with col3:
        st.metric("Total Excess", f"{data['total_excess']:.2f} kWh")
    with col4:
        st.metric("Status", data['excess_status'])
    
    # Additional details
    st.subheader("📈 Data Summary")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.info(f"**Unique Days (Generated):** {data['unique_days_gen']}")
    with col2:
        st.info(f"**Unique Days (Consumed):** {data['unique_days_cons']}")
    with col3:
        st.info(f"**Total Generated (Before Loss):** {data['sum_injection']:.2f} kWh")
    
    # Financial Calculations Display on Web Page
    st.subheader("💰 Financial Calculations")
    
    # Helper function for consistent rounding throughout the application
    def round_kwh_financial(value):
        return int(value + 0.5) if value >= 0 else int(value - 0.5)
    
    # Check if financial calculation data is available
    if 'total_excess_financial_rounded' in data:
        col1, col2 = st.columns(2)

        total_excess_financial_rounded = data['total_excess_financial_rounded']
        additional_surcharge_value = data.get('additional_surcharge', 0.0)
        additional_surcharge_rate = data.get('additional_surcharge_rate', 0.0)
        additional_surcharge_period_label = data.get('additional_surcharge_period_label') or "Selected Period"
        additional_surcharge_note = data.get('additional_surcharge_note', '')
        additional_surcharge_kwh = data.get('additional_surcharge_kwh', data.get('iex_excess_financial', 0))

        tariff_label = data.get('tariff_selection', TARIFF_OPTIONS[0])
        tariff_window_label = data.get('tariff_window_label', '')
        tariff_c1_c2_rate = data.get('tariff_c1_c2_rate', data.get('tariff_rates', {}).get('c1_c2_rate', 0))
        tariff_c5_rate = data.get('tariff_c5_rate', data.get('tariff_rates', {}).get('c5_rate', 0))
        tariff_cross_subsidy_rate = data.get('tariff_cross_subsidy_rate', data.get('tariff_rates', {}).get('cross_subsidy_rate', 0))
        tariff_wheeling_rate = data.get('tariff_wheeling_rate', data.get('tariff_rates', {}).get('wheeling_rate', 0.0))
        wheeling_reference = data.get('wheeling_reference_kwh', 0.0)
        wheeling_combined = data.get('wheeling_combined_kwh', total_excess_financial_rounded + wheeling_reference)
        wheeling_reduction = data.get('wheeling_reduction_kwh', wheeling_combined * WHEELING_PERCENT)
        wheeling_adjusted = data.get('wheeling_adjusted_kwh', wheeling_combined - wheeling_reduction)

        st.info(f"**Tariff Applied:** {tariff_label} ({tariff_window_label or 'Latest Tariff'})")

        with col1:
            st.write("**Positive Charges:**")
            st.info(f"**Base Rate:** {data['total_excess_financial_rounded']} kWh × Rs.{data['base_rate']:.4f} = Rs.{data['base_amount']:.2f}")
            st.info(f"**C1+C2 Additional:** {data['c1_c2_excess']} kWh × Rs.{tariff_c1_c2_rate:.4f} = Rs.{data['c1_c2_additional']:.2f}")
            st.info(f"**C5 Additional:** {data['c5_excess']} kWh × Rs.{tariff_c5_rate:.4f} = Rs.{data['c5_additional']:.2f}")
            st.info(f"**Subtotal:** Rs.{data['total_amount']:.2f}")
            st.info(f"**E-Tax (5%):** Rs.{data['etax']:.2f}")
            st.success(f"**Total with E-Tax:** Rs.{data['total_with_etax']:.2f}")

        with col2:
            st.write("**Negative Charges (Deductions):**")
            st.warning(f"**E-Tax on IEX:** Rs.{data['etax_on_iex']:.2f}")
            st.warning(f"**Cross Subsidy Surcharge:** {data['iex_excess_financial']} kWh × Rs.{tariff_cross_subsidy_rate:.4f} = Rs.{data['cross_subsidy_surcharge']:.2f}")
            st.warning(
                f"**Wheeling Charges:** ({wheeling_combined:.2f} - {wheeling_reduction:.2f}) kWh × Rs.{tariff_wheeling_rate:.4f} = Rs.{data['wheeling_charges']:.2f}"
                f" (Loss add {wheeling_reference:.2f} kWh, T&D {data.get('t_and_d_loss', 0)}%)"
            )
            if additional_surcharge_value > 0:
                st.warning(
                    f"**Additional Surcharge (IEX):** {additional_surcharge_kwh} kWh × Rs.{additional_surcharge_rate:.2f}"
                    f" = Rs.{additional_surcharge_value:.2f} ({additional_surcharge_period_label})"
                )
            else:
                st.info(
                    f"**Additional Surcharge (IEX):** Not applied. {additional_surcharge_note or 'Select a month & year covered by a TNERC window.'}"
                )

            total_deductions = data['etax_on_iex'] + data['cross_subsidy_surcharge'] + data['wheeling_charges'] + additional_surcharge_value
            st.warning(f"**Total Deductions:** Rs.{total_deductions:.2f}")
            st.success(f"**Final Amount:** Rs.{data['final_amount']:.2f}")
            st.success(f"**Final Amount (Rounded Up):** Rs.{data['final_amount_rounded']}")
    else:
        # Fallback to calculating on the fly if the pre-calculated values aren't available
        # Calculate financial values using rounded values for consistency
        total_excess_financial_rounded = round_kwh_financial(data['total_excess'])

        fallback_tariff = data.get('tariff_rates') or resolve_tariff_rates(
            data.get('tariff_selection', TARIFF_OPTIONS[0]),
            data.get('month'),
            data.get('year'),
        )

        base_rate = fallback_tariff['base_rate']
        c1_c2_rate = fallback_tariff['c1_c2_rate']
        c5_rate = fallback_tariff['c5_rate']
        cross_subsidy_rate = fallback_tariff['cross_subsidy_rate']
        wheeling_rate = fallback_tariff['wheeling_rate']
        base_amount = total_excess_financial_rounded * base_rate

        # Calculate TOD-wise excess for financial calculations
        merged_data = data.get('merged_all', pd.DataFrame())
        additional_surcharge = 0.0
        additional_surcharge_breakdown = []
        additional_surcharge_rate = 0.0
        additional_surcharge_period_label = "Month & Year not selected"
        additional_surcharge_note = "Select a valid month & year to apply Additional Surcharge"
        if not merged_data.empty:
            tod_excess = merged_data.groupby('TOD_Category')['Total_Excess'].sum().reset_index()

            # Additional charges for specific TOD categories using rounded values
            c1_c2_excess_raw = tod_excess.loc[tod_excess['TOD_Category'].isin(['C1', 'C2']), 'Total_Excess'].sum()
            c1_c2_excess = round_kwh_financial(c1_c2_excess_raw)
            c1_c2_additional = c1_c2_excess * c1_c2_rate  # rupees per kWh

            c5_excess_raw = tod_excess.loc[tod_excess['TOD_Category'] == 'C5', 'Total_Excess'].sum()
            c5_excess = round_kwh_financial(c5_excess_raw)
            c5_additional = c5_excess * c5_rate  # rupees per kWh

            # Calculate total amount
            total_amount = base_amount + c1_c2_additional + c5_additional

            # Calculate E-Tax (5% of total amount)
            etax = total_amount * 0.05

            # Calculate total amount with E-Tax
            total_with_etax = total_amount + etax

            # Calculate IEX excess for specific charges using rounded values
            iex_excess_financial_raw = merged_data['IEX_Excess'].sum() if 'IEX_Excess' in merged_data.columns else data['total_excess']
            iex_excess_financial = round_kwh_financial(iex_excess_financial_raw)

            # Calculate negative factors using rounded values
            etax_on_iex = total_excess_financial_rounded * 0.1
            cross_subsidy_surcharge = iex_excess_financial * cross_subsidy_rate  # Only for IEX excess
            additional_surcharge, additional_surcharge_breakdown, additional_surcharge_rate, additional_surcharge_period_label, additional_surcharge_note = calculate_monthly_additional_surcharge(
                data.get('month'),
                data.get('year'),
                iex_excess_financial_raw,
                iex_excess_financial,
            )

            # Wheeling charges derived from total excess and configured T&D loss
            wheeling_components = compute_wheeling_components(
                total_excess_financial_rounded,
                data.get('t_and_d_loss', 0.0),
            )
            wheeling_reference_kwh = wheeling_components['reference_kwh']
            wheeling_combined_kwh = wheeling_components['combined_kwh']
            wheeling_reduction_kwh = wheeling_components['reduction_kwh']
            wheeling_adjusted_kwh = wheeling_components['adjusted_kwh']
            wheeling_charges = wheeling_adjusted_kwh * wheeling_rate

            # Calculate final amount to be collected (include Additional Surcharge as deduction)
            final_amount = total_with_etax - (etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge)

            # Round up final amount to next highest value
            final_amount_rounded = math.ceil(final_amount)

            # Display financial calculations in organized columns
            col1, col2 = st.columns(2)

            with col1:
                st.write("**Positive Charges:**")
                st.info(f"**Base Rate:** {total_excess_financial_rounded} kWh × Rs.{base_rate:.4f} = Rs.{base_amount:.2f}")
                st.info(f"**C1+C2 Additional:** {c1_c2_excess} kWh × Rs.{c1_c2_rate:.4f} = Rs.{c1_c2_additional:.2f}")
                st.info(f"**C5 Additional:** {c5_excess} kWh × Rs.{c5_rate:.4f} = Rs.{c5_additional:.2f}")
                st.info(f"**Subtotal:** Rs.{total_amount:.2f}")
                st.info(f"**E-Tax (5%):** Rs.{etax:.2f}")
                st.success(f"**Total with E-Tax:** Rs.{total_with_etax:.2f}")

            with col2:
                st.write("**Negative Charges (Deductions):**")
                st.warning(f"**E-Tax on IEX:** Rs.{etax_on_iex:.2f}")
                st.warning(f"**Cross Subsidy Surcharge:** {iex_excess_financial} kWh × Rs.{cross_subsidy_rate:.4f} = Rs.{cross_subsidy_surcharge:.2f}")
                st.warning(
                    f"**Wheeling Charges:** ({wheeling_combined_kwh:.2f} - {wheeling_reduction_kwh:.2f}) kWh × Rs.{wheeling_rate:.4f} = Rs.{wheeling_charges:.2f}"
                    f" (Loss add {wheeling_reference_kwh:.2f} kWh, T&D {data.get('t_and_d_loss', 0)}%)"
                )
                if additional_surcharge > 0:
                    st.warning(
                        f"**Additional Surcharge (IEX):** {iex_excess_financial} kWh × Rs.{additional_surcharge_rate:.2f}"
                        f" = Rs.{additional_surcharge:.2f} ({additional_surcharge_period_label})"
                    )
                else:
                    st.info(
                        f"**Additional Surcharge (IEX):** Not applied. {additional_surcharge_note or 'Select a month & year covered by a TNERC window.'}"
                    )
                total_deductions = etax_on_iex + cross_subsidy_surcharge + wheeling_charges + additional_surcharge
                st.warning(f"**Total Deductions:** Rs.{total_deductions:.2f}")
                st.success(f"**Final Amount:** Rs.{final_amount:.2f}")
                st.success(f"**Final Amount (Rounded Up):** Rs.{final_amount_rounded}")
        else:
            st.warning("No merged data available for financial calculations.")
    
    # TOD-wise breakdown
    st.subheader("⏰ TOD-wise Excess Energy Breakdown")
    merged_data = data.get('merged_all', pd.DataFrame())
    if not merged_data.empty:
        tod_excess = merged_data.groupby('TOD_Category')['Total_Excess'].sum().reset_index()
        tod_display = []
        for _, row in tod_excess.iterrows():
            category = row['TOD_Category']
            excess_rounded = round_kwh_financial(row['Total_Excess'])
            tod_display.append({"TOD Category": category, "Excess Energy (kWh)": excess_rounded})
        
        if tod_display:
            tod_df = pd.DataFrame(tod_display)
            st.dataframe(tod_df, use_container_width=True)
    else:
        st.warning("No TOD data available for breakdown.")
    
    # Generate PDFs based on selected options from top
    st.header("📄 Generating PDF Reports")
    st.info("Processing your data and generating PDF reports based on your selections...")
    
    with st.spinner("Generating PDF reports..."):
        # Generate PDFs based on checkbox selections
        pdfs_generated = []
        
        if show_excess_only:
            if not data['merged_excess'].empty:
                pdf_bytes = generate_detailed_pdf(data, data['merged_excess'], "excess")
                if pdf_bytes:
                    filename = generate_custom_filename("excess_only", data['consumer_number'], data['consumer_name'], data.get('month'), data.get('year'))
                    pdfs_generated.append((filename, pdf_bytes))
        
        if show_all_slots:
            pdf_bytes = generate_detailed_pdf(data, data['merged_all'], "all_slots")
            if pdf_bytes:
                filename = generate_custom_filename("all_slots", data['consumer_number'], data['consumer_name'], data.get('month'), data.get('year'))
                pdfs_generated.append((filename, pdf_bytes))
        
        if show_daywise:
            pdf_bytes = generate_daywise_pdf(data, data['merged_all'])
            if pdf_bytes:
                filename = generate_custom_filename("daywise", data['consumer_number'], data['consumer_name'], data.get('month'), data.get('year'))
                pdfs_generated.append((filename, pdf_bytes))
    
    # Display download option (complete package only)
    if pdfs_generated:
        st.success(f"✅ Successfully generated {len(pdfs_generated)} PDF report(s)!")

        excel_artifacts_raw = data.get('uploaded_artifacts') or []
        excel_downloads = []
        if excel_artifacts_raw:
            for idx, artifact in enumerate(excel_artifacts_raw, start=1):
                extension = artifact.get('extension') or '.xlsx'
                generated_name = generate_custom_filename(
                    artifact.get('base_name', f"source_{idx}"),
                    data['consumer_number'],
                    data['consumer_name'],
                    data.get('month'),
                    data.get('year'),
                    extension=extension,
                )
                hashed_name = append_hash_suffix(generated_name, artifact.get('hash'))
                excel_downloads.append({
                    'final_name': hashed_name,
                    'bytes': artifact['bytes'],
                    'hash_full': artifact.get('hash', ''),
                    'hash_short': (artifact.get('hash', '') or '')[:12],
                    'original_name': artifact.get('original_name', 'uploaded_file.xlsx'),
                })

        if excel_downloads:
            st.subheader("🛡️ Source Excel Integrity Pack")
            integrity_rows = [
                {
                    "Renamed File": entry['final_name'],
                    "Original Upload": entry['original_name'],
                    "SHA256 (first 12 chars)": entry['hash_short'] or 'N/A',
                }
                for entry in excel_downloads
            ]
            if integrity_rows:
                integrity_df = pd.DataFrame(integrity_rows)
                st.dataframe(integrity_df, use_container_width=True)

        slug = build_consumer_slug(data['consumer_number'], data['consumer_name'])
        bundle_buffer = io.BytesIO()
        manifest_lines = [
            "Complete Package Manifest",
            "-------------------------",
        ]
        with zipfile.ZipFile(bundle_buffer, 'w') as bundle_zip:
            if pdfs_generated:
                manifest_lines.append("Reports Included:")
                for fname, pdf_bytes in pdfs_generated:
                    bundle_zip.writestr(f"reports/{fname}", pdf_bytes)
                    manifest_lines.append(f" - {fname}")
            if excel_downloads:
                manifest_lines.append("")
                manifest_lines.append("Source Excels:")
                for entry in excel_downloads:
                    bundle_zip.writestr(f"source_excels/{entry['final_name']}", entry['bytes'])
                    manifest_lines.append(
                        f" - {entry['final_name']} | sha256={entry['hash_full']} | original={entry['original_name']}"
                    )
            bundle_zip.writestr("MANIFEST.txt", "\n".join(manifest_lines))
        bundle_buffer.seek(0)
        bundle_zip_name = f"{slug}_reports_and_sources.zip"
        st.download_button(
            label="📦 Download Complete Package (ZIP)",
            data=bundle_buffer.getvalue(),
            file_name=bundle_zip_name,
            mime="application/zip",
            type="primary"
        )
    else:
        st.error("❌ No PDF reports were generated. Please check your selections and try again.")

# Display errors
if st.session_state.error_message:
    st.error(st.session_state.error_message)

render_auto_update_sidebar()
render_footer("Energy Adjustment Calculator - Streamlit Version")
